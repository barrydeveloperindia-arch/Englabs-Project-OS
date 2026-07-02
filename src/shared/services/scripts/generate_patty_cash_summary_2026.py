import os
import json
import openpyxl
import copy
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference, PieChart
from collections import defaultdict
import datetime

def get_client_name(project_id):
    project_id_clean = str(project_id).strip()
    if not project_id_clean or project_id_clean == 'C-XXXX' or project_id_clean == 'CXXXX':
        return "Non-Project / General Expense"
    
    local_path = f"data/{project_id_clean}.json"
    if os.path.exists(local_path):
        try:
            with open(local_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data.get('client', 'Unknown')
        except Exception:
            pass
            
    known_clients = {
        'C5023': 'AV Machining',
        'C5195': 'Labat Asia',
        'C5254': 'Sonalika',
        'C5237': 'Bajaj',
        'C5230': 'Enertics',
        'C5283': 'AutoDyanamics',
        'C5223': 'Henkel',
        'C5239': 'SARK',
        'C5244': 'N/A',
        'C5008': 'Henkel',
        'C5247': 'N/A',
        'C5182': 'N/A',
        'C5285': 'AutoDyanamics'
    }
    return known_clients.get(project_id_clean, "Unknown Client")

def safe_copy_styles(src_cell, dest_cell):
    if src_cell.font:
        dest_cell.font = copy.copy(src_cell.font)
    if src_cell.fill:
        dest_cell.fill = copy.copy(src_cell.fill)
    if src_cell.border:
        dest_cell.border = copy.copy(src_cell.border)
    if src_cell.alignment:
        dest_cell.alignment = copy.copy(src_cell.alignment)
    if src_cell.number_format:
        dest_cell.number_format = src_cell.number_format

def main():
    source_excel = 'G:\\HR Team Managements\\Englabs Projects APK\\Patty Cash Details\\Englabs Patty Cash.xlsx'
    output_dir = 'G:\\HR Team Managements\\Englabs Projects APK\\Patty Cash Details'
    output_excel = os.path.join(output_dir, 'Patty Cash Summary 2026.xlsx')
    
    print(f"Loading source data workbook: {source_excel}")
    wb_data = openpyxl.load_workbook(source_excel, data_only=True)
    print("Loading source formula workbook...")
    wb_form = openpyxl.load_workbook(source_excel, data_only=False)
    
    months = ['Jan_2026', 'Feb_2026', 'Mar_2026', 'Apr_2026', 'May_2026', 'Jun_2026', 'July_2026']
    
    all_transactions = []
    monthly_summaries = {}
    project_monthly_expenses = defaultdict(lambda: defaultdict(float))
    all_project_ids = set()
    
    for name in months:
        if name not in wb_data.sheetnames:
            continue
            
        sheet_d = wb_data[name]
        sheet_f = wb_form[name]
        
        header_row = None
        for r in range(1, 10):
            val = sheet_d.cell(row=r, column=1).value
            if val == 'Date':
                header_row = r
                break
        if not header_row:
            for r in range(1, 10):
                row_vals = [sheet_d.cell(row=r, column=c).value for c in range(1, 10)]
                if 'Date' in row_vals or 'Detail Of Transition / Expenses' in row_vals:
                    header_row = r
                    break
        if not header_row:
            continue
            
        cat_row = header_row - 1 if header_row > 1 else header_row
        
        totals_rows = set()
        for r in range(1, header_row + 3):
            cell_f_f = sheet_f.cell(row=r, column=6).value
            cell_f_g = sheet_f.cell(row=r, column=7).value
            for cell_val in [cell_f_f, cell_f_g]:
                if cell_val and isinstance(cell_val, str) and cell_val.strip().startswith('='):
                    totals_rows.add(r)
                    
        bottom_table_start = sheet_d.max_row + 1
        for r in range(header_row + 2, sheet_d.max_row + 1):
            for c in range(1, sheet_d.max_column + 1):
                val = sheet_d.cell(row=r, column=c).value
                if val and any(k in str(val).lower() for k in ['projects wise details', 'project details', 'closing may-2026', 'april-2026 projects']):
                    bottom_table_start = r
                    break
            if bottom_table_start <= sheet_d.max_row:
                break
                
        proj_id_col = None
        proj_dr_col = None
        category_cols = {}
        
        categories_list = ['Maintenance', 'Emergency', 'Zepto / Blinkit', 'Sky5 Hotel', 'Mr. Behl', 'Gurpreet Porter', 'Project Dr.']
        
        for col_idx in range(1, sheet_d.max_column + 1):
            c_val = sheet_d.cell(row=cat_row, column=col_idx).value
            h_val = sheet_d.cell(row=header_row, column=col_idx).value
            
            for val in [c_val, h_val]:
                if val:
                    val_str = str(val).lower()
                    if 'project id' in val_str:
                        proj_id_col = col_idx
                    elif 'project dr' in val_str:
                        proj_dr_col = col_idx
                    
                    for cat in categories_list:
                        if cat.lower() in val_str:
                            category_cols[col_idx] = cat
                            
        if not proj_id_col:
            for r in [header_row - 1, header_row, header_row + 1]:
                if r > 0:
                    for col_idx in range(1, sheet_d.max_column + 1):
                        val = sheet_d.cell(row=r, column=col_idx).value
                        if val and 'c-xxxx' in str(val).lower():
                            proj_id_col = col_idx
                            break
                            
        start_row = header_row + 1
        
        month_dr = 0.0
        month_cr = 0.0
        month_cats = defaultdict(float)
        
        for r in range(start_row, bottom_table_start):
            if r in totals_rows:
                continue
                
            desc = sheet_d.cell(row=r, column=5).value
            dr = sheet_d.cell(row=r, column=7).value
            cr = sheet_d.cell(row=r, column=6).value
            date_val = sheet_d.cell(row=r, column=1).value
            
            if not desc and not dr and not cr and not date_val:
                continue
            if desc and any(k in str(desc).lower() for k in ['total', 'grand total', 'balance b/f', 'balance c/f']):
                continue
                
            dr_val = 0.0
            cr_val = 0.0
            if dr is not None:
                try: dr_val = float(dr)
                except ValueError: pass
            if cr is not None:
                try: cr_val = float(cr)
                except ValueError: pass
                    
            month_dr += dr_val
            month_cr += cr_val
            
            row_cats = {}
            for col_idx, cat_name in category_cols.items():
                cell_val = sheet_d.cell(row=r, column=col_idx).value
                if cell_val is not None:
                    try:
                        val_float = float(cell_val)
                        row_cats[cat_name] = val_float
                        month_cats[cat_name] += val_float
                    except ValueError:
                        pass
                        
            pid_val = ""
            if proj_id_col:
                pid = sheet_d.cell(row=r, column=proj_id_col).value
                if pid:
                    pid_val = str(pid).strip()
                    all_project_ids.add(pid_val)
                    
            proj_dr_val = row_cats.get('Project Dr.', 0.0)
            if pid_val and proj_dr_val > 0:
                project_monthly_expenses[pid_val][name] += proj_dr_val
                
            formatted_date = ""
            if date_val:
                if isinstance(date_val, datetime.datetime):
                    formatted_date = date_val.strftime("%Y-%m-%d")
                else:
                    formatted_date = str(date_val)
                    
            all_transactions.append({
                'month': name,
                'date': formatted_date,
                'from': sheet_d.cell(row=r, column=2).value or "",
                'paymentMode': sheet_d.cell(row=r, column=3).value or "",
                'receivedBy': sheet_d.cell(row=r, column=4).value or "",
                'details': desc or "",
                'cr': cr_val,
                'dr': dr_val,
                'categories': row_cats,
                'projectId': pid_val,
                'remarks': sheet_d.cell(row=r, column=sheet_d.max_column).value if sheet_d.max_column > 17 else ""
            })
            
        monthly_summaries[name] = {
            'cr': month_cr,
            'dr': month_dr,
            'balance': month_cr - month_dr,
            'categories': month_cats
        }
        
    # Create new workbook
    new_wb = openpyxl.Workbook()
    
    # Styles
    navy_fill = PatternFill(start_color='0E4368', end_color='0E4368', fill_type='solid')
    light_blue_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    light_green_fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    title_font = Font(name='Segoe UI', size=16, bold=True, color='0E4368')
    bold_font = Font(name='Segoe UI', size=10, bold=True)
    normal_font = Font(name='Segoe UI', size=10)
    kpi_title_font = Font(name='Segoe UI', size=9, color='64748B', bold=True)
    kpi_value_font = Font(name='Segoe UI', size=16, bold=True, color='0E4368')
    kpi_green_font = Font(name='Segoe UI', size=16, bold=True, color='059669')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    border_thin = Side(style='thin', color='CBD5E1')
    border_double = Side(style='double', color='0E4368')
    grid_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    total_border = Border(top=border_thin, bottom=border_double)
    
    # Sheet 1: Consolidated_Overview
    ws_ov = new_wb.active
    ws_ov.title = "Consolidated_Overview"
    ws_ov.views.sheetView[0].showGridLines = True
    
    ws_ov['A1'] = "Englabs Petty Cash 2026 - Consolidated Overview"
    ws_ov['A1'].font = title_font
    
    # ── KPI Cards Setup ──
    # Card 1: Total Credit
    ws_ov.merge_cells("A3:C3")
    ws_ov["A3"] = "TOTAL CREDIT RECEIVED"
    ws_ov["A3"].font = kpi_title_font
    ws_ov["A3"].alignment = center_align
    ws_ov["A3"].fill = light_blue_fill
    ws_ov.merge_cells("A4:C5")
    ws_ov["A4"] = "=B16"  # Link to Total Credit in Table Row 16
    ws_ov["A4"].font = kpi_value_font
    ws_ov["A4"].alignment = center_align
    ws_ov["A4"].number_format = '₹#,##0.00'
    ws_ov["A4"].fill = light_blue_fill
    
    # Card 2: Total Debit
    ws_ov.merge_cells("D3:F3")
    ws_ov["D3"] = "TOTAL DEBIT OUTFLOW"
    ws_ov["D3"].font = kpi_title_font
    ws_ov["D3"].alignment = center_align
    ws_ov["D3"].fill = light_blue_fill
    ws_ov.merge_cells("D4:F5")
    ws_ov["D4"] = "=C16"  # Link to Total Debit in Table Row 16
    ws_ov["D4"].font = kpi_value_font
    ws_ov["D4"].alignment = center_align
    ws_ov["D4"].number_format = '₹#,##0.00'
    ws_ov["D4"].fill = light_blue_fill
    
    # Card 3: Net Balance
    ws_ov.merge_cells("G3:I3")
    ws_ov["G3"] = "CURRENT NET BALANCE"
    ws_ov["G3"].font = kpi_title_font
    ws_ov["G3"].alignment = center_align
    ws_ov["G3"].fill = light_green_fill
    ws_ov.merge_cells("G4:I5")
    ws_ov["G4"] = "=D16"  # Link to Total Balance in Table Row 16
    ws_ov["G4"].font = kpi_green_font
    ws_ov["G4"].alignment = center_align
    ws_ov["G4"].number_format = '₹#,##0.00'
    ws_ov["G4"].fill = light_green_fill
    
    # Add border to KPI cards
    for r in range(3, 6):
        for c in range(1, 10):
            ws_ov.cell(row=r, column=c).border = grid_border
            
    # Overview Table Headers (Row 8)
    ov_headers = ["Month", "Credit (Cr.)", "Debit (Dr.)", "Net Balance", "Maintenance", "Emergency", "Zepto / Blinkit", "Sky5 Hotel", "Mr. Behl", "Gurpreet Porter", "Project Dr."]
    for col_idx, h in enumerate(ov_headers, 1):
        cell = ws_ov.cell(row=8, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = center_align
        
    current_row = 9
    for m_name in months:
        summary = monthly_summaries.get(m_name, {'cr':0.0, 'dr':0.0, 'balance':0.0, 'categories':{}})
        ws_ov.cell(row=current_row, column=1, value=m_name.replace('_2026', '')).font = bold_font
        
        ws_ov.cell(row=current_row, column=2, value=summary['cr']).number_format = '₹#,##0.00'
        ws_ov.cell(row=current_row, column=3, value=summary['dr']).number_format = '₹#,##0.00'
        ws_ov.cell(row=current_row, column=4, value=summary['balance']).number_format = '₹#,##0.00'
        
        cats_map = ['Maintenance', 'Emergency', 'Zepto / Blinkit', 'Sky5 Hotel', 'Mr. Behl', 'Gurpreet Porter', 'Project Dr.']
        for col_idx, cat in enumerate(cats_map, 5):
            amt = summary['categories'].get(cat, 0.0)
            ws_ov.cell(row=current_row, column=col_idx, value=amt).number_format = '₹#,##0.00'
            
        for c in range(1, 12):
            cell = ws_ov.cell(row=current_row, column=c)
            if c > 1:
                cell.font = normal_font
                cell.alignment = right_align
            cell.border = grid_border
            
        current_row += 1
        
    # Totals row (Row 16)
    ws_ov.cell(row=current_row, column=1, value="Total Summary").font = bold_font
    ws_ov.cell(row=current_row, column=1).alignment = left_align
    ws_ov.cell(row=current_row, column=1).border = total_border
    
    for c in range(2, 12):
        col_let = get_column_letter(c)
        cell = ws_ov.cell(row=current_row, column=c, value=f"=SUM({col_let}9:{col_let}{current_row-1})")
        cell.font = bold_font
        cell.alignment = right_align
        cell.number_format = '₹#,##0.00'
        cell.border = total_border
        
    for col_idx in range(1, 12):
        ws_ov.column_dimensions[get_column_letter(col_idx)].width = 18
        
    # ── Add Charts ──
    # Column Chart: Credit vs Debit
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Monthly Cash Flow (Credit vs Debit)"
    chart.y_axis.title = "Amount (Rs.)"
    chart.x_axis.title = "Month"
    chart.width = 15
    chart.height = 10
    
    data = Reference(ws_ov, min_col=2, min_row=8, max_col=3, max_row=15)
    cats = Reference(ws_ov, min_col=1, min_row=9, max_row=15)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_ov.add_chart(chart, "A19")
    
    # Pie Chart: Category distribution
    pie = PieChart()
    pie.title = "Expense Distribution by Category"
    pie.width = 15
    pie.height = 10
    
    labels = Reference(ws_ov, min_col=5, min_row=8, max_col=11, max_row=8)
    data_pie = Reference(ws_ov, min_col=5, min_row=16, max_col=11, max_row=16) # Sum totals row
    pie.add_data(data_pie, from_rows=True, titles_from_data=False)
    pie.set_categories(labels)
    ws_ov.add_chart(pie, "G19")
    
    # Sheet 2: Project_Summary
    ws_proj = new_wb.create_sheet(title="Project_Summary")
    ws_proj.views.sheetView[0].showGridLines = True
    ws_proj['A1'] = "Consolidated Project-wise Petty Cash Expenses 2026"
    ws_proj['A1'].font = title_font
    
    # Project summary headers
    proj_headers = ["Project ID", "Client Name", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "July", "Cumulative Total"]
    for col_idx, h in enumerate(proj_headers, 1):
        cell = ws_proj.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = center_align
        
    current_row = 4
    project_cumulatives = []
    for pid in all_project_ids:
        total_p = sum(project_monthly_expenses[pid][m] for m in months)
        project_cumulatives.append((pid, total_p))
        
    project_cumulatives.sort(key=lambda x: x[1], reverse=True)
    
    for pid, cumul in project_cumulatives:
        ws_proj.cell(row=current_row, column=1, value=pid).font = bold_font
        ws_proj.cell(row=current_row, column=2, value=get_client_name(pid)).font = normal_font
        
        for col_idx, m_name in enumerate(months, 3):
            val = project_monthly_expenses[pid][m_name]
            ws_proj.cell(row=current_row, column=col_idx, value=val if val > 0 else "").number_format = '₹#,##0.00'
            
        ws_proj.cell(row=current_row, column=10, value=f"=SUM(C{current_row}:I{current_row})").number_format = '₹#,##0.00'
        ws_proj.cell(row=current_row, column=10).font = bold_font
        
        for c in range(1, 11):
            cell = ws_proj.cell(row=current_row, column=c)
            cell.border = grid_border
            if c > 2:
                cell.alignment = right_align
                
        current_row += 1
        
    ws_proj.cell(row=current_row, column=1, value="Total").font = bold_font
    ws_proj.cell(row=current_row, column=1).border = total_border
    ws_proj.cell(row=current_row, column=2, value="").border = total_border
    
    for c in range(3, 11):
        col_let = get_column_letter(c)
        cell = ws_proj.cell(row=current_row, column=c, value=f"=SUM({col_let}4:{col_let}{current_row-1})")
        cell.font = bold_font
        cell.alignment = right_align
        cell.number_format = '₹#,##0.00'
        cell.border = total_border
        
    ws_proj.column_dimensions['A'].width = 15
    ws_proj.column_dimensions['B'].width = 30
    for c in range(3, 11):
        ws_proj.column_dimensions[get_column_letter(c)].width = 15
        
    # Sheet 3: Project_Filter_Lookup
    ws_look = new_wb.create_sheet(title="Project_Filter_Lookup")
    ws_look.views.sheetView[0].showGridLines = True
    
    ws_look['A1'] = "Project ID Transaction Lookup Center"
    ws_look['A1'].font = title_font
    
    ws_look['A2'] = "Select Project ID:"
    ws_look['A2'].font = bold_font
    
    pids_sorted = sorted(list(all_project_ids))
    for i, pid in enumerate(pids_sorted, 1):
        ws_look.cell(row=i, column=26, value=pid)
        
    dv = DataValidation(type="list", formula1=f"=Z1:Z{len(pids_sorted)}", allow_blank=True)
    dv.error ='Your entry is not in the project list'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please select from the list'
    dv.promptTitle = 'Project ID List'
    
    ws_look.add_data_validation(dv)
    dv.add(ws_look['B2'])
    
    ws_look['B2'].font = Font(name='Segoe UI', size=11, bold=True, color='0E4368')
    ws_look['B2'].border = Border(left=Side(style='medium', color='0E4368'), right=Side(style='medium', color='0E4368'),
                                   top=Side(style='medium', color='0E4368'), bottom=Side(style='medium', color='0E4368'))
    ws_look['B2'].fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    ws_look['B2'].alignment = center_align
    
    lookup_headers = ["Month", "Date", "From", "Payment Mode", "Received By", "Transaction Details", "Credit (Cr.)", "Debit (Dr.)", 
                      "Maintenance", "Emergency", "Zepto / Blinkit", "Sky5 Hotel", "Mr. Behl", "Gurpreet Porter", "Project Dr.", "Remarks"]
                      
    for col_idx, h in enumerate(lookup_headers, 1):
        cell = ws_look.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = center_align
        
    formula_filter = f'=FILTER(CHOOSECOLS(All_Transactions!A4:Q{len(all_transactions)+10}, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 17), All_Transactions!P4:P{len(all_transactions)+10} = B2, "No transactions found")'
    ws_look.cell(row=5, column=1, value=formula_filter)
    
    for c in range(1, 17):
        ws_look.column_dimensions[get_column_letter(c)].width = 18
        
    # Sheet 4: All_Transactions
    ws_all = new_wb.create_sheet(title="All_Transactions")
    ws_all.views.sheetView[0].showGridLines = True
    ws_all['A1'] = "All Petty Cash Transactions 2026 (Date-wise)"
    ws_all['A1'].font = title_font
    
    all_headers = ["Month", "Date", "From", "Payment Mode", "Received By", "Transaction Details", "Credit (Cr.)", "Debit (Dr.)", 
                   "Maintenance", "Emergency", "Zepto / Blinkit", "Sky5 Hotel", "Mr. Behl", "Gurpreet Porter", "Project Dr.", "Project ID", "Remarks"]
                   
    for col_idx, h in enumerate(all_headers, 1):
        cell = ws_all.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = center_align
        
    current_row = 4
    for idx, tx in enumerate(all_transactions, 4):
        ws_all.cell(row=current_row, column=1, value=tx['month'].replace('_2026', '')).font = normal_font
        ws_all.cell(row=current_row, column=2, value=tx['date']).font = normal_font
        ws_all.cell(row=current_row, column=3, value=tx['from']).font = normal_font
        ws_all.cell(row=current_row, column=4, value=tx['paymentMode']).font = normal_font
        ws_all.cell(row=current_row, column=5, value=tx['receivedBy']).font = normal_font
        ws_all.cell(row=current_row, column=6, value=tx['details']).font = normal_font
        ws_all.cell(row=current_row, column=7, value=tx['cr'] if tx['cr'] > 0 else "").number_format = '₹#,##0.00'
        ws_all.cell(row=current_row, column=8, value=tx['dr'] if tx['dr'] > 0 else "").number_format = '₹#,##0.00'
        
        cats = ['Maintenance', 'Emergency', 'Zepto / Blinkit', 'Sky5 Hotel', 'Mr. Behl', 'Gurpreet Porter', 'Project Dr.']
        for col_idx, cat in enumerate(cats, 9):
            val = tx['categories'].get(cat, 0.0)
            ws_all.cell(row=current_row, column=col_idx, value=val if val > 0 else "").number_format = '₹#,##0.00'
            
        ws_all.cell(row=current_row, column=16, value=tx['projectId']).font = bold_font
        ws_all.cell(row=current_row, column=17, value=tx['remarks']).font = normal_font
        
        for c in range(1, 18):
            cell = ws_all.cell(row=current_row, column=c)
            cell.border = grid_border
            if c in [7, 8, 9, 10, 11, 12, 13, 14, 15]:
                cell.alignment = right_align
            elif c in [1, 2, 4, 16]:
                cell.alignment = center_align
                
        current_row += 1
        
    for c in range(1, 18):
        max_len = max(len(str(ws_all.cell(row=r, column=c).value or '')) for r in range(3, min(current_row, 30)))
        ws_all.column_dimensions[get_column_letter(c)].width = max(max_len + 3, 15)

    # 5. Copy individual month sheets exactly as tabs
    for name in months:
        if name not in wb_form.sheetnames:
            continue
            
        print(f"Copying monthly sheet tab: {name}")
        src_sheet = wb_form[name]
        dest_sheet = new_wb.create_sheet(title=name)
        dest_sheet.views.sheetView[0].showGridLines = True
        
        for row in src_sheet.iter_rows():
            for cell in row:
                dest_cell = dest_sheet.cell(row=cell.row, column=cell.column)
                dest_cell.value = cell.value
                safe_copy_styles(cell, dest_cell)
                
        for merged_range in src_sheet.merged_cells.ranges:
            try:
                dest_sheet.merge_cells(str(merged_range))
            except Exception:
                pass
                
        for col, dim in src_sheet.column_dimensions.items():
            if dim.width:
                dest_sheet.column_dimensions[col].width = dim.width
                
    new_wb.save(output_excel)
    print(f"SUCCESS: Premium consolidated workbook generated at: {output_excel}")

if __name__ == '__main__':
    main()
