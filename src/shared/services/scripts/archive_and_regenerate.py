import os
import shutil
import datetime

# Directories
base_dir = 'G:\\HR Team Managements\\Englabs Projects APK\\Store\\Rate List Record'
target_dir = os.path.join(base_dir, 'Rate List')
old_dir = os.path.join(target_dir, 'OLD')

# Ensure OLD directory exists
if not os.path.exists(old_dir):
    os.makedirs(old_dir)
    print("Created OLD directory:", old_dir)

# Timestamp for archiving
timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

# Files to archive
files_to_archive = [
    'Englabs_Store_Rate_Registry_2026.xlsx',
    'Englabs_Store_Rate_Registry_Report.pdf',
    'Englabs_Store_Rate_Registry_Report_Detailed.pdf'
]

for filename in files_to_archive:
    src_path = os.path.join(target_dir, filename)
    if os.path.exists(src_path):
        # Generate archived name with timestamp
        base, ext = os.path.splitext(filename)
        dest_filename = f"{base}_{timestamp}{ext}"
        dest_path = os.path.join(old_dir, dest_filename)
        
        try:
            shutil.move(src_path, dest_path)
            print(f"Archived {filename} -> OLD/{dest_filename}")
        except Exception as e:
            print(f"Failed to archive {filename}:", e)

# Now regenerate the coded excel and pdf
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

json_path = 'c:/Users/SAM/Documents/Antigravity/Englabs Projects/scratch/compiled_detailed_list_coded.json'
logo_path = 'c:\\Users\\SAM\\Documents\\Antigravity\\Englabs Projects\\src\\assets\\englabs_logo.png'

output_excel = os.path.join(target_dir, 'Englabs_Store_Rate_Registry_2026.xlsx')
output_pdf = os.path.join(target_dir, 'Englabs_Store_Rate_Registry_Report_Detailed.pdf')

# Load compiled list
with open(json_path, 'r', encoding='utf-8') as f:
    compiled_list = json.load(f)

# Sort alphabetically by name
compiled_list = sorted(compiled_list, key=lambda x: x['code'])

# ==========================================
# WRITE EXCEL WORKBOOK
# ==========================================
wb = openpyxl.Workbook()

# Sheet 1: Dashboard
ws_dash = wb.active
ws_dash.title = "Rate_Dashboard"
ws_dash.views.sheetView[0].showGridLines = True

navy_fill = PatternFill(start_color="092A42", end_color="092A42", fill_type="solid")
emerald_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
card_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")

title_font = Font(name="Calibri", size=18, bold=True, color="092A42")
subtitle_font = Font(name="Calibri", size=11, italic=True, color="64748B")
card_title_font = Font(name="Calibri", size=10, bold=True, color="94A3B8")
card_val_font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")

ws_dash['A1'] = "ENGLABS STORE RATE REGISTRY"
ws_dash['A1'].font = title_font
ws_dash['A2'] = "Corporate Inventory Performance & Procurement Rate List Record (Live 2026)"
ws_dash['A2'].font = subtitle_font

ws_dash.merge_cells('A4:B4')
ws_dash['A4'] = "TOTAL STORE ITEMS"
ws_dash['A4'].font = card_title_font
ws_dash['A4'].alignment = Alignment(horizontal="center")
ws_dash.merge_cells('A5:B5')
ws_dash['A5'] = len(compiled_list)
ws_dash['A5'].font = card_val_font
ws_dash['A5'].alignment = Alignment(horizontal="center")

ws_dash.merge_cells('C4:D4')
ws_dash['C4'] = "GENERAL ITEMS"
ws_dash['C4'].font = card_title_font
ws_dash['C4'].alignment = Alignment(horizontal="center")
ws_dash.merge_cells('C5:D5')
ws_dash['C5'] = len([i for i in compiled_list if i['category'].lower() in ['general', 'stationery', 'paints', 'chemicals']])
ws_dash['C5'].font = card_val_font
ws_dash['C5'].alignment = Alignment(horizontal="center")

ws_dash.merge_cells('E4:F4')
ws_dash['E4'] = "RAW MATERIALS"
ws_dash['E4'].font = card_title_font
ws_dash['E4'].alignment = Alignment(horizontal="center")
ws_dash.merge_cells('E5:F5')
ws_dash['E5'] = len([i for i in compiled_list if i['category'].lower() not in ['general', 'stationery', 'paints', 'chemicals']])
ws_dash['E5'].font = card_val_font
ws_dash['E5'].alignment = Alignment(horizontal="center")

for row in [4, 5]:
    for col in range(1, 7):
        cell = ws_dash.cell(row=row, column=col)
        cell.fill = card_fill

# Sheet 2: Store_Rate_Registry
ws_reg = wb.create_sheet(title="Store_Rate_Registry")
ws_reg.views.sheetView[0].showGridLines = True

headers = [
    "S.No.", "Material Code", "Material Name", "Category", "Rack Location",
    "Current Stock", "Unit", "Min Threshold", "Latest Inward Date",
    "Current Rate (₹)", "Previous Rate (₹)", "Rate Change (₹)", "Change %", "Remarks / Source"
]

for col_idx, h in enumerate(headers, 1):
    cell = ws_reg.cell(row=1, column=col_idx)
    cell.value = h
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws_reg.row_dimensions[1].height = 28

thin_border = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

for r_idx, item in enumerate(compiled_list, 2):
    ws_reg.cell(row=r_idx, column=1, value=r_idx - 1).alignment = Alignment(horizontal="center")
    ws_reg.cell(row=r_idx, column=2, value=item['code']).alignment = Alignment(horizontal="center")
    ws_reg.cell(row=r_idx, column=3, value=item['name']).alignment = Alignment(horizontal="left")
    ws_reg.cell(row=r_idx, column=4, value=item['category']).alignment = Alignment(horizontal="center")
    ws_reg.cell(row=r_idx, column=5, value=item['qty']).alignment = Alignment(horizontal="center")
    ws_reg.cell(row=r_idx, column=6, value=item['qty']).alignment = Alignment(horizontal="center")
    ws_reg.cell(row=r_idx, column=7, value=item['unit']).alignment = Alignment(horizontal="center")
    ws_reg.cell(row=r_idx, column=8, value=item['min_q']).alignment = Alignment(horizontal="center")
    ws_reg.cell(row=r_idx, column=9, value=item['inward_date']).alignment = Alignment(horizontal="center")
    
    c_cell = ws_reg.cell(row=r_idx, column=10, value=item['c_rate'])
    c_cell.number_format = '₹#,##0.00'
    c_cell.alignment = Alignment(horizontal="right")
    
    p_cell = ws_reg.cell(row=r_idx, column=11, value=item['p_rate'])
    p_cell.number_format = '₹#,##0.00'
    p_cell.alignment = Alignment(horizontal="right")
    
    ch_cell = ws_reg.cell(row=r_idx, column=12, value=item['change'])
    ch_cell.number_format = '₹#,##0.00'
    ch_cell.alignment = Alignment(horizontal="right")
    
    pct_cell = ws_reg.cell(row=r_idx, column=13, value=item['change_pct']/100.0)
    pct_cell.number_format = '0.00%'
    pct_cell.alignment = Alignment(horizontal="right")
    
    if item['change'] > 0:
        ch_cell.font = Font(name="Calibri", size=10, bold=True, color="DC2626")
        pct_cell.font = Font(name="Calibri", size=10, bold=True, color="DC2626")
    elif item['change'] < 0:
        ch_cell.font = Font(name="Calibri", size=10, bold=True, color="16A34A")
        pct_cell.font = Font(name="Calibri", size=10, bold=True, color="16A34A")
        
    ws_reg.cell(row=r_idx, column=14, value=item['note']).alignment = Alignment(horizontal="left")
    
    bg_color = "F8FAFC" if r_idx % 2 == 0 else "FFFFFF"
    fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    for col_idx in range(1, len(headers) + 1):
        c = ws_reg.cell(row=r_idx, column=col_idx)
        c.border = thin_border
        if col_idx not in [12, 13] or item['change'] == 0:
            c.fill = fill

for col in ws_reg.columns:
    max_len = 0
    col_let = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            if cell.column in [10, 11, 12] and isinstance(cell.value, (int, float)):
                val_len = len(f"₹{cell.value:,.2f}")
            else:
                val_len = len(str(cell.value))
            if val_len > max_len:
                max_len = val_len
    ws_reg.column_dimensions[col_let].width = max(max_len + 3, 10)

ws_reg.column_dimensions['C'].width = 35
ws_reg.column_dimensions['N'].width = 35

wb.save(output_excel)
print("SUCCESS: Englabs_Store_Rate_Registry_2026.xlsx updated.")

# ==========================================
# WRITE REPORTLAB PDF
# ==========================================
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica-Bold", 8)
        self.setFillColor(colors.HexColor("#092A42"))
        
        # Header
        self.setStrokeColor(colors.HexColor("#E2E8F0"))
        self.setLineWidth(0.5)
        self.line(36, A4[1] - 40, A4[0] - 36, A4[1] - 40)
        self.drawString(36, A4[1] - 35, "ENGLABS INDIA PVT. LTD. — STORE RATE REGISTRY REPORT")
        
        # Footer
        self.line(36, 45, A4[0] - 36, 45)
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(A4[0] - 36, 32, page_text)
        self.drawString(36, 32, "Confidential — Internal Store Operations")
        self.restoreState()

def build_pdf():
    doc = SimpleDocTemplate(
        output_pdf,
        pagesize=A4,
        leftMargin=36,
        rightMargin=36,
        topMargin=54,
        bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#092A42'),
        spaceAfter=2
    )
    
    meta_style = ParagraphStyle(
        'MetaText',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=15
    )
    
    section_style = ParagraphStyle(
        'SecTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=16,
        textColor=colors.HexColor('#092A42'),
        spaceBefore=15,
        spaceAfter=8
    )
    
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7,
        leading=9,
        textColor=colors.HexColor('#1E293B')
    )
    
    cell_style_bold = ParagraphStyle(
        'CellTextBold',
        parent=cell_style,
        fontName='Helvetica-Bold'
    )
    
    cell_style_center = ParagraphStyle(
        'CellTextCenter',
        parent=cell_style,
        alignment=1
    )
    
    cell_style_right = ParagraphStyle(
        'CellTextRight',
        parent=cell_style,
        alignment=2
    )
    
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=colors.white,
        alignment=1
    )
    
    header_style_left = ParagraphStyle(
        'HeaderStyleLeft',
        parent=header_style,
        alignment=0
    )

    story = []
    
    # Logo
    logo_w = 40
    logo_h = 40
    if os.path.exists(logo_path):
        logo_img = Image(logo_path, width=logo_w, height=logo_h)
        logo_img.hAlign = 'LEFT'
        story.append(logo_img)
        story.append(Spacer(1, 5))
        
    story.append(Paragraph("Englabs Store Rate Registry Report", title_style))
    story.append(Paragraph("Detailed Performance, Rack Locations & Procurement Dates Diagnostics (Live 2026)", meta_style))
    
    # KPI Grid
    gen_cnt = len([i for i in compiled_list if i['category'].lower() in ['general', 'stationery', 'paints', 'chemicals']])
    raw_cnt = len([i for i in compiled_list if i['category'].lower() not in ['general', 'stationery', 'paints', 'chemicals']])
    
    kpi_title_style = ParagraphStyle('KPITitle', parent=cell_style_center, fontName='Helvetica-Bold', fontSize=8, textColor=colors.HexColor('#475569'))
    kpi_val_style_navy = ParagraphStyle('KPIValNavy', parent=cell_style_center, fontName='Helvetica-Bold', fontSize=18, leading=20, textColor=colors.HexColor('#092A42'))
    kpi_val_style_emerald = ParagraphStyle('KPIValEmer', parent=cell_style_center, fontName='Helvetica-Bold', fontSize=18, leading=20, textColor=colors.HexColor('#10B981'))
    kpi_val_style_blue = ParagraphStyle('KPIValBlue', parent=cell_style_center, fontName='Helvetica-Bold', fontSize=18, leading=20, textColor=colors.HexColor('#0284C7'))
    
    kpi_data = [
        [
            Paragraph("TOTAL UNIQUE ITEMS", kpi_title_style),
            Paragraph("GENERAL CONSUMABLES", kpi_title_style),
            Paragraph("RAW MATERIALS", kpi_title_style),
        ],
        [
            Paragraph("304", kpi_val_style_navy),
            Paragraph(str(gen_cnt), kpi_val_style_emerald),
            Paragraph(str(raw_cnt), kpi_val_style_blue),
        ]
    ]
    kpi_table = Table(kpi_data, colWidths=[174, 174, 174])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#E2E8F0')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 10))
    
    story.append(Paragraph("Master Store Item Registry", section_style))
    
    # Table headers
    table_data = [[
        Paragraph("<b>S.No.</b>", header_style),
        Paragraph("<b>Material Code</b>", header_style),
        Paragraph("<b>Material Name</b>", header_style_left),
        Paragraph("<b>Rack</b>", header_style),
        Paragraph("<b>Qty / Unit</b>", header_style),
        Paragraph("<b>Inward Date</b>", header_style),
        Paragraph("<b>Current Rate</b>", header_style),
        Paragraph("<b>Prev Rate</b>", header_style),
        Paragraph("<b>Rate Change</b>", header_style),
    ]]
    
    for idx, item in enumerate(compiled_list, 1):
        c_rate_str = f"Rs. {item['c_rate']:.2f}"
        p_rate_str = f"Rs. {item['p_rate']:.2f}"
        
        change_val = item['change']
        if change_val > 0:
            change_str = f"+Rs. {change_val:.2f}"
            change_color = '#DC2626'
        elif change_val < 0:
            change_str = f"-Rs. {abs(change_val):.2f}"
            change_color = '#16A34A'
        else:
            change_str = "Rs. 0.00"
            change_color = '#475569'
            
        change_para = Paragraph(f"<font color='{change_color}'><b>{change_str}</b></font>", cell_style_right)
        
        table_data.append([
            Paragraph(str(idx), cell_style_center),
            Paragraph(item['code'], cell_style_center),
            Paragraph(item['name'], cell_style),
            Paragraph(item['location'], cell_style_center),
            Paragraph(f"{item['qty']}", cell_style_center),
            Paragraph(item['inward_date'], cell_style_center),
            Paragraph(c_rate_str, cell_style_right),
            Paragraph(p_rate_str, cell_style_right),
            change_para,
        ])
        
    col_widths = [20, 50, 153, 50, 55, 55, 45, 45, 50]
    
    t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#092A42')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ])
    
    for i in range(1, len(table_data)):
        bg = colors.HexColor('#F8FAFC') if i % 2 == 0 else colors.white
        t_style.add('BACKGROUND', (0, i), (-1, i), bg)
        t_style.add('TOPPADDING', (0, i), (-1, i), 3)
        t_style.add('BOTTOMPADDING', (0, i), (-1, i), 3)
        
    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    main_table.setStyle(t_style)
    story.append(main_table)
    
    doc.build(story, canvasmaker=NumberedCanvas)
    print("SUCCESS: Englabs_Store_Rate_Registry_Report_Detailed.pdf updated.")

build_pdf()
