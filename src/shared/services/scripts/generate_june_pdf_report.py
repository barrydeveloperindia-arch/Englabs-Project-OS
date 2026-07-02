import os
import json
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, KeepTogether, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from datetime import datetime
from collections import defaultdict

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            # Skip page numbers on cover page (Page 1)
            if self._pageNumber > 1:
                self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#64748B"))
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawCentredString(A4[0] / 2.0, 20, page_text)
        print_date = datetime.now().strftime("%d-%b-%Y %I:%M %p")
        self.drawRightString(A4[0] - 35, 20, f"Generated: {print_date} • Englabs Projects")
        self.drawString(35, 20, "Confidential - June 2026 Petty Cash Ledger")
        self.restoreState()

def get_client_name(project_id):
    project_id_clean = str(project_id).strip()
    if not project_id_clean or project_id_clean == 'C-XXXX' or project_id_clean == 'CXXXX':
        return "Non-Project / General Expense"
    
    local_path = f"data/{project_id_clean}.json"
    if os.path.exists(local_path):
        try:
            with open(local_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data.get('client', 'Unknown')
        except Exception:
            pass
            
    known_clients = {
        'C5023': 'AV Machining',
        'C5195': 'Labat Asia',
        'C5254': 'Sonalika',
        'C5237': 'Bajaj',
        'C5230': 'Enertics',
        'C5283': 'AutoDyanamics',
        'C5223': 'Henkel',
        'C5239': 'SARK',
        'C5244': 'N/A',
        'C5008': 'Henkel',
        'C5247': 'N/A',
        'C5182': 'N/A',
        'C5285': 'AutoDyanamics'
    }
    return known_clients.get(project_id_clean, "Unknown Client")

def main():
    source_excel = 'G:\\HR Team Managements\\Englabs Projects APK\\Patty Cash Details\\Englabs Patty Cash.xlsx'
    output_dir = 'G:\\HR Team Managements\\Englabs Projects APK\\Patty Cash Details\\June_2026_Summary_Report'
    output_pdf = os.path.join(output_dir, 'June_2026_Detailed_Petty_Cash_Report.pdf')
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    print("Loading data workbook...")
    wb = openpyxl.load_workbook(source_excel, data_only=True)
    sheet = wb['Jun_2026']
    
    # 1. Parse June 2026 details
    category_totals = {
        'Maintenance': 0.0,
        'Emergency': 0.0,
        'Zepto / Blinkit': 0.0,
        'Sky5 Hotel': 0.0,
        'Mr. Behl': 0.0,
        'Gurpreet Porter': 0.0,
        'Project Dr.': 0.0
    }
    
    categories_cols = {
        'J': 'Maintenance',
        'K': 'Emergency',
        'L': 'Zepto / Blinkit',
        'M': 'Sky5 Hotel',
        'N': 'Mr. Behl',
        'O': 'Gurpreet Porter',
        'P': 'Project Dr.'
    }
    
    tx_list = []
    project_totals = defaultdict(float)
    
    total_cr = 0.0
    total_dr = 0.0
    
    for r in range(5, 137): # June transaction rows
        desc = sheet.cell(row=r, column=5).value
        dr = sheet.cell(row=r, column=7).value
        cr = sheet.cell(row=r, column=6).value
        date_val = sheet.cell(row=r, column=1).value
        proj_id = sheet.cell(row=r, column=17).value
        proj_dr = sheet.cell(row=r, column=16).value
        
        date_str = ""
        if date_val:
            if isinstance(date_val, datetime):
                date_str = date_val.strftime("%d-%b")
            elif isinstance(date_val, datetime.date):
                date_str = date_val.strftime("%d-%b")
            else:
                date_str = str(date_val)
                
        desc_str = ""
        if desc:
            try:
                desc_str = str(desc).strip().encode('ascii', 'ignore').decode('ascii')
            except Exception:
                desc_str = str(desc).strip()
                
        dr_val = float(dr) if dr is not None else 0.0
        cr_val = float(cr) if cr is not None else 0.0
        pid_str = str(proj_id).strip() if proj_id else ""
        pdr_val = float(proj_dr) if proj_dr is not None else 0.0
        
        total_cr += cr_val
        total_dr += dr_val
        
        for col_let, cat_name in categories_cols.items():
            cell_val = sheet[f"{col_let}{r}"].value
            if cell_val is not None:
                try:
                    category_totals[cat_name] += float(cell_val)
                except ValueError:
                    pass
                    
        if pid_str:
            project_totals[pid_str] += pdr_val
            
        tx_list.append({
            'date': date_str,
            'desc': desc_str,
            'cr': cr_val,
            'dr': dr_val,
            'pid': pid_str,
            'pdr': pdr_val
        })
        
    # Get courier budgets mapping for project-wise section
    courier_budgets = {
        'C5023': (2500, 2500),
        'C5195': (1200, 0),
        'C5254': (1000, 1000),
        'C5239': (0, 0),
        'C5251': (6000, 5000),
        'C5280': (1000, 1000),
        'C5267': (1000, 1000),
        'C5295': (1000, 2100),
        'C5264': (0, 8500),
        'C5255': (700, 700),
        'C5287': (1000, 200),
        'C5310': (1000, 1000),
        'C5286': (2500, 1000)
    }
    
    # 2. Build PDF report
    doc = SimpleDocTemplate(
        output_pdf,
        pagesize=A4,
        leftMargin=35,
        rightMargin=35,
        topMargin=35,
        bottomMargin=40
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        name='TitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=28,
        textColor=colors.HexColor('#0E4368'),
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        name='SubTitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=15,
        textColor=colors.HexColor('#059669'),
        spaceAfter=30
    )
    section_title_style = ParagraphStyle(
        name='SectionTitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=11,
        leading=14,
        textColor=colors.HexColor('#0E4368'),
        spaceBefore=14,
        spaceAfter=8
    )
    body_style = ParagraphStyle(
        name='BodyStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9.5,
        leading=13,
        textColor=colors.HexColor('#475569')
    )
    meta_style = ParagraphStyle(
        name='MetaStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#1E293B')
    )
    cell_normal_style = ParagraphStyle(
        name='CellNormalStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7,
        leading=9,
        textColor=colors.HexColor('#334155')
    )
    cell_bold_style = ParagraphStyle(
        name='CellBoldStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7,
        leading=9,
        textColor=colors.HexColor('#1E293B')
    )
    cell_red_bold_style = ParagraphStyle(
        name='CellRedBoldStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7,
        leading=9,
        textColor=colors.HexColor('#DC2626')
    )
    cell_green_bold_style = ParagraphStyle(
        name='CellGreenBoldStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7,
        leading=9,
        textColor=colors.HexColor('#16A34A')
    )
    cell_header_style = ParagraphStyle(
        name='CellHeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7.5,
        leading=10,
        textColor=colors.white
    )
    
    # ── Page 1: EXECUTIVE COVER PAGE ──
    story.append(Spacer(1, 40))
    logo_file = "src/assets/englabs_logo.png"
    if os.path.exists(logo_file):
        story.append(RLImage(logo_file, width=65, height=65))
        story.append(Spacer(1, 15))
        
    story.append(Paragraph("ENGLABS PROJECTS", title_style))
    story.append(Paragraph("PETTY CASH AUDIT & COURIER PERFORMANCE REPORT", subtitle_style))
    
    # Metadata block
    meta_data = [
        [Paragraph("Reporting Period:", cell_bold_style), Paragraph("June 2026", body_style)],
        [Paragraph("Document Type:", cell_bold_style), Paragraph("Monthly Financial Performance & Budget Variance Audit", body_style)],
        [Paragraph("Company Name:", cell_bold_style), Paragraph("Englabs India Pvt. Ltd", body_style)],
        [Paragraph("Audit Date:", cell_bold_style), Paragraph(datetime.now().strftime("%B %d, %Y"), body_style)],
        [Paragraph("Classification:", cell_bold_style), Paragraph("Confidential / Management Review", body_style)]
    ]
    meta_table = Table(meta_data, colWidths=[120, 300])
    meta_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 40))
    
    summary_para = """
    <b>Executive Summary:</b><br/>
    This publication-quality report outlines the consolidated financial statements and transaction audit log of Englabs Projects' petty cash ledger for June 2026. This report delivers strategic business insights by cross-referencing actual project-specific expenditures with predefined client courier budgets (Vendor to Englabs and Englabs to Client). Underperforming project nodes and cost over-runs are flagged to facilitate cost controls and inventory management.
    """
    story.append(Paragraph(summary_para, body_style))
    story.append(PageBreak())
    
    # ── Page 2: Summary Dashboard ──
    story.append(Paragraph("I. Cash Outflow & Category Summary", section_title_style))
    
    # Outflow totals Table
    outflow_data = [
        [Paragraph("Category Name", cell_header_style), Paragraph("Total Spent (₹)", cell_header_style)],
        [Paragraph("Maintenance", cell_normal_style), Paragraph(f"₹ {category_totals['Maintenance']:,.2f}", cell_normal_style)],
        [Paragraph("Emergency", cell_normal_style), Paragraph(f"₹ {category_totals['Emergency']:,.2f}", cell_normal_style)],
        [Paragraph("Zepto / Blinkit", cell_normal_style), Paragraph(f"₹ {category_totals['Zepto / Blinkit']:,.2f}", cell_normal_style)],
        [Paragraph("Sky5 Hotel", cell_normal_style), Paragraph(f"₹ {category_totals['Sky5 Hotel']:,.2f}", cell_normal_style)],
        [Paragraph("Mr. Behl", cell_normal_style), Paragraph(f"₹ {category_totals['Mr. Behl']:,.2f}", cell_normal_style)],
        [Paragraph("Gurpreet Porter", cell_normal_style), Paragraph(f"₹ {category_totals['Gurpreet Porter']:,.2f}", cell_normal_style)],
        [Paragraph("Project Dr. (Direct Project Expenses)", cell_normal_style), Paragraph(f"₹ {category_totals['Project Dr.']:,.2f}", cell_normal_style)],
        [Paragraph("Total Outflow (Debit)", cell_bold_style), Paragraph(f"₹ {total_dr:,.2f}", cell_bold_style)],
        [Paragraph("Total Credit (Cr.)", cell_bold_style), Paragraph(f"₹ {total_cr:,.2f}", cell_bold_style)],
        [Paragraph("Net Balance", cell_bold_style), Paragraph(f"₹ {total_cr - total_dr:,.2f}", cell_bold_style)]
    ]
    
    outflow_table = Table(outflow_data, colWidths=[280, 245])
    t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0E4368')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 5),
        ('TOPPADDING', (0,0), (-1,0), 5),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BOX', (0,0), (-1,-1), 0.8, colors.HexColor('#0E4368')),
    ])
    for idx in range(1, len(outflow_data) - 3):
        bg = colors.HexColor('#F8FAFC') if idx % 2 == 0 else colors.white
        t_style.add('BACKGROUND', (0, idx), (-1, idx), bg)
        t_style.add('BOTTOMPADDING', (0, idx), (-1, idx), 4)
        t_style.add('TOPPADDING', (0, idx), (-1, idx), 4)
    t_style.add('BACKGROUND', (0, -3), (-1, -1), colors.HexColor('#E2E8F0'))
    t_style.add('BOTTOMPADDING', (0, -3), (-1, -1), 5)
    t_style.add('TOPPADDING', (0, -3), (-1, -1), 5)
    outflow_table.setStyle(t_style)
    story.append(outflow_table)
    story.append(Spacer(1, 10))
    
    # II. Project summary vs budgets table
    story.append(Paragraph("II. June Project Payments vs Courier Budgets", section_title_style))
    
    proj_headers = [
        Paragraph("Project ID", cell_header_style),
        Paragraph("Client Name", cell_header_style),
        Paragraph("Actual Dr. (₹)", cell_header_style),
        Paragraph("VND ➔ ENG Budget (₹)", cell_header_style),
        Paragraph("ENG ➔ CLT Budget (₹)", cell_header_style)
    ]
    proj_rows = [proj_headers]
    
    sorted_pids = sorted(project_totals.items(), key=lambda x: x[1], reverse=True)
    for pid, tot in sorted_pids:
        bud_v2e, bud_e2c = courier_budgets.get(pid, (None, None))
        v2e_str = f"₹{bud_v2e:,.2f}" if bud_v2e is not None else "-"
        e2c_str = f"₹{bud_e2c:,.2f}" if bud_e2c is not None else "-"
        
        proj_rows.append([
            Paragraph(pid, cell_bold_style),
            Paragraph(get_client_name(pid), cell_normal_style),
            Paragraph(f"₹{tot:,.2f}", cell_normal_style),
            Paragraph(v2e_str, cell_normal_style),
            Paragraph(e2c_str, cell_normal_style)
        ])
        
    proj_table = Table(proj_rows, colWidths=[75, 170, 90, 95, 95])
    proj_t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0E4368')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 5),
        ('TOPPADDING', (0,0), (-1,0), 5),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BOX', (0,0), (-1,-1), 0.8, colors.HexColor('#0E4368')),
    ])
    for idx in range(1, len(proj_rows)):
        bg = colors.HexColor('#F8FAFC') if idx % 2 == 0 else colors.white
        proj_t_style.add('BACKGROUND', (0, idx), (-1, idx), bg)
        proj_t_style.add('BOTTOMPADDING', (0, idx), (-1, idx), 4)
        proj_t_style.add('TOPPADDING', (0, idx), (-1, idx), 4)
    proj_table.setStyle(proj_t_style)
    story.append(proj_table)
    story.append(PageBreak())
    
    # ── Page 3: Budget Variance Analysis ──
    story.append(Paragraph("III. Project Courier Budget Variance Analysis", section_title_style))
    story.append(Paragraph("This section highlights projects that have exceeded their total assigned courier budgets.", body_style))
    story.append(Spacer(1, 10))
    
    var_headers = [
        Paragraph("Project ID", cell_header_style),
        Paragraph("Client Name", cell_header_style),
        Paragraph("Total Budget (₹)", cell_header_style),
        Paragraph("Actual Dr. (₹)", cell_header_style),
        Paragraph("Variance (₹)", cell_header_style),
        Paragraph("Status", cell_header_style)
    ]
    var_rows = [var_headers]
    
    for pid, tot in sorted_pids:
        bud_v2e, bud_e2c = courier_budgets.get(pid, (None, None))
        if bud_v2e is None and bud_e2c is None:
            continue  # Only evaluate projects with budget
            
        total_budget = (bud_v2e or 0) + (bud_e2c or 0)
        variance = total_budget - tot
        
        if variance < 0:
            status_para = Paragraph("<b>🔴 OVER BUDGET</b>", cell_red_bold_style)
            var_str = f"-₹{abs(variance):,.2f}"
            var_para = Paragraph(var_str, cell_red_bold_style)
        else:
            status_para = Paragraph("<b>🟢 UNDER BUDGET</b>", cell_green_bold_style)
            var_str = f"+₹{variance:,.2f}"
            var_para = Paragraph(var_str, cell_green_bold_style)
            
        var_rows.append([
            Paragraph(pid, cell_bold_style),
            Paragraph(get_client_name(pid), cell_normal_style),
            Paragraph(f"₹{total_budget:,.2f}", cell_normal_style),
            Paragraph(f"₹{tot:,.2f}", cell_normal_style),
            var_para,
            status_para
        ])
        
    var_table = Table(var_rows, colWidths=[70, 150, 95, 75, 80, 55])
    var_t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0E4368')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 5),
        ('TOPPADDING', (0,0), (-1,0), 5),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BOX', (0,0), (-1,-1), 0.8, colors.HexColor('#0E4368')),
    ])
    for idx in range(1, len(var_rows)):
        bg = colors.HexColor('#F8FAFC') if idx % 2 == 0 else colors.white
        var_t_style.add('BACKGROUND', (0, idx), (-1, idx), bg)
        var_t_style.add('BOTTOMPADDING', (0, idx), (-1, idx), 5)
        var_t_style.add('TOPPADDING', (0, idx), (-1, idx), 5)
    var_table.setStyle(var_t_style)
    story.append(var_table)
    story.append(PageBreak())
    
    # ── Page 4+: Date-wise Transaction Details ──
    story.append(Paragraph("IV. Date-wise Detailed Transactions Log", section_title_style))
    
    log_headers = [
        Paragraph("Date", cell_header_style),
        Paragraph("Transaction Description", cell_header_style),
        Paragraph("Cr. (₹)", cell_header_style),
        Paragraph("Dr. (₹)", cell_header_style),
        Paragraph("Proj ID", cell_header_style),
        Paragraph("Proj Dr. (₹)", cell_header_style)
    ]
    log_rows = [log_headers]
    
    for tx in tx_list:
        cr_str = f"₹{tx['cr']:,.2f}" if tx['cr'] > 0 else "-"
        dr_str = f"₹{tx['dr']:,.2f}" if tx['dr'] > 0 else "-"
        pdr_str = f"₹{tx['pdr']:,.2f}" if tx['pdr'] > 0 else "-"
        pid_str = tx['pid'] if tx['pid'] else "-"
        
        log_rows.append([
            Paragraph(tx['date'], cell_normal_style),
            Paragraph(tx['desc'], cell_normal_style),
            Paragraph(cr_str, cell_normal_style),
            Paragraph(dr_str, cell_normal_style),
            Paragraph(pid_str, cell_bold_style if tx['pid'] else cell_normal_style),
            Paragraph(pdr_str, cell_normal_style)
        ])
        
    log_table = Table(log_rows, colWidths=[45, 230, 60, 60, 65, 65], repeatRows=1)
    log_t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0E4368')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 4),
        ('TOPPADDING', (0,0), (-1,0), 4),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BOX', (0,0), (-1,-1), 0.8, colors.HexColor('#0E4368')),
    ])
    for idx in range(1, len(log_rows)):
        bg = colors.HexColor('#F8FAFC') if idx % 2 == 0 else colors.white
        log_t_style.add('BACKGROUND', (0, idx), (-1, idx), bg)
        log_t_style.add('BOTTOMPADDING', (0, idx), (-1, idx), 3)
        log_t_style.add('TOPPADDING', (0, idx), (-1, idx), 3)
    log_table.setStyle(log_t_style)
    story.append(log_table)
    
    doc.build(story, canvasmaker=NumberedCanvas)
    print("SUCCESS: Premium June report PDF generated.")

if __name__ == '__main__':
    main()
