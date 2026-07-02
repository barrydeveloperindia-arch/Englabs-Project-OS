import os
import json
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, KeepTogether, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from datetime import datetime
from collections import defaultdict

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            if self._pageNumber > 1:
                self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#64748B"))
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawCentredString(A4[0] / 2.0, 20, page_text)
        print_date = datetime.now().strftime("%d-%b-%Y %I:%M %p")
        self.drawRightString(A4[0] - 35, 20, f"Generated: {print_date} • Englabs Projects")
        self.drawString(35, 20, "Confidential - 2026 Consolidated Petty Cash Report")
        self.restoreState()

def get_client_name(project_id):
    project_id_clean = str(project_id).strip()
    if not project_id_clean or project_id_clean == 'C-XXXX' or project_id_clean == 'CXXXX':
        return "Non-Project / General Expense"
    
    local_path = f"data/{project_id_clean}.json"
    if os.path.exists(local_path):
        try:
            with open(local_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data.get('client', 'Unknown')
        except Exception:
            pass
            
    known_clients = {
        'C5023': 'AV Machining',
        'C5195': 'Labat Asia',
        'C5254': 'Sonalika',
        'C5237': 'Bajaj',
        'C5230': 'Enertics',
        'C5283': 'AutoDyanamics',
        'C5223': 'Henkel',
        'C5239': 'SARK',
        'C5244': 'N/A',
        'C5008': 'Henkel',
        'C5247': 'N/A',
        'C5182': 'N/A',
        'C5285': 'AutoDyanamics',
        'C5206': 'Hella',
        'C5238': 'Akash Bhardwaj'
    }
    return known_clients.get(project_id_clean, "Unknown Client")

def main():
    source_excel = 'G:\\HR Team Managements\\Englabs Projects APK\\Patty Cash Details\\Englabs Patty Cash.xlsx'
    output_dir = 'G:\\HR Team Managements\\Englabs Projects APK\\Patty Cash Details\\June_2026_Summary_Report'
    output_pdf = os.path.join(output_dir, 'Englabs_2026_Consolidated_Financial_Report.pdf')
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    print("Loading data workbook...")
    wb = openpyxl.load_workbook(source_excel, data_only=True)
    
    months = ['Jan_2026', 'Feb_2026', 'Mar_2026', 'Apr_2026', 'May_2026', 'Jun_2026', 'July_2026']
    
    monthly_summaries = {}
    project_totals_june = defaultdict(float)
    project_totals_july = defaultdict(float)
    
    cumulative_cr = 0.0
    cumulative_dr = 0.0
    
    cumulative_categories = {
        'Maintenance': 0.0,
        'Emergency': 0.0,
        'Zepto / Blinkit': 0.0,
        'Sky5 Hotel': 0.0,
        'Mr. Behl': 0.0,
        'Gurpreet Porter': 0.0,
        'Project Dr.': 0.0
    }
    
    # Category columns differ slightly between sheets, let's parse each sheet dynamically
    for name in months:
        if name not in wb.sheetnames:
            continue
            
        sheet = wb[name]
        
        # Detect header row
        header_row = None
        for r in range(1, 10):
            val = sheet.cell(row=r, column=1).value
            if val == 'Date':
                header_row = r
                break
        if not header_row:
            for r in range(1, 10):
                row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 10)]
                if 'Date' in row_vals or 'Detail Of Transition / Expenses' in row_vals:
                    header_row = r
                    break
        if not header_row:
            continue
            
        cat_row = header_row - 1 if header_row > 1 else header_row
        
        # Categories mapping
        category_cols = {}
        categories_list = ['Maintenance', 'Emergency', 'Zepto / Blinkit', 'Sky5 Hotel', 'Mr. Behl', 'Gurpreet Porter', 'Project Dr.']
        proj_id_col = None
        
        for col_idx in range(1, sheet.max_column + 1):
            c_val = sheet.cell(row=cat_row, column=col_idx).value
            h_val = sheet.cell(row=header_row, column=col_idx).value
            for val in [c_val, h_val]:
                if val:
                    val_str = str(val).lower()
                    if 'project id' in val_str:
                        proj_id_col = col_idx
                    for cat in categories_list:
                        if cat.lower() in val_str:
                            category_cols[col_idx] = cat
                            
        if not proj_id_col:
            for r in [header_row - 1, header_row, header_row + 1]:
                if r > 0:
                    for col_idx in range(1, sheet.max_column + 1):
                        val = sheet.cell(row=r, column=col_idx).value
                        if val and 'c-xxxx' in str(val).lower():
                            proj_id_col = col_idx
                            break
                            
        # Determine table boundaries
        totals_rows = set()
        for r in range(1, header_row + 3):
            cell_val_f = sheet.cell(row=r, column=6).value
            cell_val_g = sheet.cell(row=r, column=7).value
            for cell_val in [cell_val_f, cell_val_g]:
                if cell_val and isinstance(cell_val, str) and cell_val.strip().startswith('='):
                    totals_rows.add(r)
                    
        bottom_table_start = sheet.max_row + 1
        for r in range(header_row + 2, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                val = sheet.cell(row=r, column=c).value
                if val and any(k in str(val).lower() for k in ['projects wise details', 'project details', 'closing may-2026', 'april-2026 projects']):
                    bottom_table_start = r
                    break
            if bottom_table_start <= sheet.max_row:
                break
                
        # Parse data
        month_dr = 0.0
        month_cr = 0.0
        month_cats = defaultdict(float)
        
        # Specific row limits for June and July to ensure exact match
        r_start = header_row + 1
        r_end = bottom_table_start
        if name == 'Jun_2026':
            r_start = 5
            r_end = 137
        elif name == 'July_2026':
            r_start = 6
            r_end = 18
            
        for r in range(r_start, r_end):
            if r in totals_rows:
                continue
                
            desc = sheet.cell(row=r, column=5).value
            dr = sheet.cell(row=r, column=7).value
            cr = sheet.cell(row=r, column=6).value
            date_val = sheet.cell(row=r, column=1).value
            
            if not desc and not dr and not cr and not date_val:
                continue
            if desc and any(k in str(desc).lower() for k in ['total', 'grand total', 'balance b/f', 'balance c/f']):
                continue
                
            dr_val = float(dr) if dr is not None else 0.0
            cr_val = float(cr) if cr is not None else 0.0
            
            month_dr += dr_val
            month_cr += cr_val
            
            # Categories
            for col_idx, cat_name in category_cols.items():
                cell_val = sheet.cell(row=r, column=col_idx).value
                if cell_val is not None:
                    try:
                        val_float = float(cell_val)
                        month_cats[cat_name] += val_float
                        cumulative_categories[cat_name] += val_float
                    except ValueError:
                        pass
                        
            # Project Dr.
            proj_dr = sheet.cell(row=r, column=16 if name == 'July_2026' else 16).value # Col P
            proj_id = sheet.cell(row=r, column=16 if name == 'July_2026' else 17).value # Col P in July, Q in June
            pid_str = str(proj_id).strip() if proj_id else ""
            pdr_val = 0.0
            if proj_dr is not None:
                try:
                    pdr_val = float(proj_dr)
                except ValueError:
                    pass
            
            if not pid_str and desc and 'C5254' in str(desc):
                pid_str = 'C5254'
                
            if pid_str:
                if name == 'Jun_2026':
                    project_totals_june[pid_str] += pdr_val
                elif name == 'July_2026':
                    project_totals_july[pid_str] += pdr_val
                    
        monthly_summaries[name] = {
            'cr': month_cr,
            'dr': month_dr,
            'balance': month_cr - month_dr,
            'categories': month_cats
        }
        
    # July 2026 balance is the current balance today
    current_net_balance = monthly_summaries['July_2026']['balance']
    
    # Summing all credits and debits to show total volume
    for name in months:
        cumulative_cr += monthly_summaries[name]['cr']
        cumulative_dr += monthly_summaries[name]['dr']
        
    courier_budgets = {
        'C5023': (2500, 2500),
        'C5195': (1200, 0),
        'C5254': (1000, 1000),
        'C5239': (0, 0),
        'C5251': (6000, 5000),
        'C5280': (1000, 1000),
        'C5267': (1000, 1000),
        'C5295': (1000, 2100),
        'C5264': (0, 8500),
        'C5255': (700, 700),
        'C5287': (1000, 200),
        'C5310': (1000, 1000),
        'C5286': (2500, 1000)
    }
    
    # ── PDF Generation ──
    doc = SimpleDocTemplate(
        output_pdf,
        pagesize=A4,
        leftMargin=35,
        rightMargin=35,
        topMargin=35,
        bottomMargin=40
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    navy_hex = '#1E293B'
    emerald_hex = '#059669'
    
    title_style = ParagraphStyle(
        name='TitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=28,
        textColor=colors.HexColor(navy_hex),
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        name='SubTitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=15,
        textColor=colors.HexColor(emerald_hex),
        spaceAfter=30
    )
    section_title_style = ParagraphStyle(
        name='SectionTitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=11,
        leading=14,
        textColor=colors.HexColor(navy_hex),
        spaceBefore=14,
        spaceAfter=8
    )
    body_style = ParagraphStyle(
        name='BodyStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9.5,
        leading=13,
        textColor=colors.HexColor('#475569')
    )
    cell_normal_style = ParagraphStyle(
        name='CellNormalStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7,
        leading=9,
        textColor=colors.HexColor('#334155')
    )
    cell_bold_style = ParagraphStyle(
        name='CellBoldStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7,
        leading=9,
        textColor=colors.HexColor('#1E293B')
    )
    cell_red_bold_style = ParagraphStyle(
        name='CellRedBoldStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7,
        leading=9,
        textColor=colors.HexColor('#DC2626')
    )
    cell_green_bold_style = ParagraphStyle(
        name='CellGreenBoldStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7,
        leading=9,
        textColor=colors.HexColor('#16A34A')
    )
    cell_header_style = ParagraphStyle(
        name='CellHeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7.5,
        leading=10,
        textColor=colors.white
    )
    
    # ── Page 1: COVER PAGE ──
    story.append(Spacer(1, 40))
    logo_file = "src/assets/englabs_logo.png"
    if os.path.exists(logo_file):
        story.append(RLImage(logo_file, width=65, height=65))
        story.append(Spacer(1, 15))
        
    story.append(Paragraph("ENGLABS PROJECTS", title_style))
    story.append(Paragraph("CONSOLIDATED PETTY CASH & COURIER AUDIT REPORT 2026", subtitle_style))
    
    meta_data = [
        [Paragraph("Reporting Period:", cell_bold_style), Paragraph("January 2026 - July 2026", body_style)],
        [Paragraph("Document Type:", cell_bold_style), Paragraph("Annual Consolidated Financial Summary & Audit Report", body_style)],
        [Paragraph("Company Name:", cell_bold_style), Paragraph("Englabs India Pvt. Ltd", body_style)],
        [Paragraph("Audit Date:", cell_bold_style), Paragraph(datetime.now().strftime("%B %d, %Y"), body_style)],
        [Paragraph("Classification:", cell_bold_style), Paragraph("Confidential / Management Review", body_style)]
    ]
    meta_table = Table(meta_data, colWidths=[120, 300])
    meta_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 40))
    
    summary_para = """
    <b>Executive Summary:</b><br/>
    This consolidated report presents the annual financial statement and transaction audit of Englabs Projects' petty cash ledger for the year 2026 (January to July). It includes high-level monthly summaries, category-wise cash distributions, and detailed budget-to-expenditure evaluations for both June and July. Key variance metrics are highlighted to assist executive decision-making.
    """
    story.append(Paragraph(summary_para, body_style))
    story.append(PageBreak())
    
    # ── Page 2: EXCEL FIRST PAGE (CONSOLIDATED OVERVIEW) ──
    story.append(Paragraph("I. 2026 Consolidated Monthly Cash Flow (Excel First Page)", section_title_style))
    
    # KPI Headers in Table
    kpi_table_data = [
        [Paragraph("CUMULATIVE CREDIT", cell_header_style), Paragraph("CUMULATIVE DEBIT", cell_header_style), Paragraph("CURRENT NET BALANCE", cell_header_style)],
        [Paragraph(f"₹ {cumulative_cr:,.2f}", cell_bold_style), Paragraph(f"₹ {cumulative_dr:,.2f}", cell_bold_style), Paragraph(f"₹ {current_net_balance:,.2f}", cell_green_bold_style)]
    ]
    kpi_table = Table(kpi_table_data, colWidths=[175, 175, 175])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor(navy_hex)),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor(navy_hex)),
        ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#F1F5F9')),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 15))
    
    # Monthly Cash Flow Table
    flow_headers = [
        Paragraph("Month", cell_header_style),
        Paragraph("Credit (Cr. - ₹)", cell_header_style),
        Paragraph("Debit (Dr. - ₹)", cell_header_style),
        Paragraph("Net Balance (₹)", cell_header_style)
    ]
    flow_rows = [flow_headers]
    for name in months:
        sum_m = monthly_summaries.get(name, {'cr':0.0, 'dr':0.0, 'balance':0.0})
        flow_rows.append([
            Paragraph(name.replace('_2026', ''), cell_bold_style),
            Paragraph(f"₹ {sum_m['cr']:,.2f}", cell_normal_style),
            Paragraph(f"₹ {sum_m['dr']:,.2f}", cell_normal_style),
            Paragraph(f"₹ {sum_m['balance']:,.2f}", cell_normal_style)
        ])
    # Total Row
    flow_rows.append([
        Paragraph("Total Summary", cell_bold_style),
        Paragraph(f"₹ {cumulative_cr:,.2f}", cell_bold_style),
        Paragraph(f"₹ {cumulative_dr:,.2f}", cell_bold_style),
        Paragraph(f"₹ {current_net_balance:,.2f}", cell_green_bold_style)
    ])
    
    flow_table = Table(flow_rows, colWidths=[110, 140, 140, 135])
    flow_t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor(navy_hex)),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BOX', (0,0), (-1,-1), 0.8, colors.HexColor(navy_hex)),
    ])
    for idx in range(1, len(flow_rows) - 1):
        bg = colors.HexColor('#F8FAFC') if idx % 2 == 0 else colors.white
        flow_t_style.add('BACKGROUND', (0, idx), (-1, idx), bg)
    flow_t_style.add('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E2E8F0'))
    flow_table.setStyle(flow_t_style)
    story.append(flow_table)
    
    # ── Category Cumulative Table ──
    story.append(Spacer(1, 15))
    story.append(Paragraph("II. 2026 Expense Distribution by Category", section_title_style))
    
    cat_headers = [Paragraph("Category Name", cell_header_style), Paragraph("Total Cumulative Outflow (₹)", cell_header_style)]
    cat_rows = [cat_headers]
    for c_name, val in cumulative_categories.items():
        cat_rows.append([
            Paragraph(c_name, cell_normal_style),
            Paragraph(f"₹ {val:,.2f}", cell_normal_style)
        ])
    cat_rows.append([
        Paragraph("Total Outflow", cell_bold_style),
        Paragraph(f"₹ {cumulative_dr:,.2f}", cell_bold_style)
    ])
    
    cat_table = Table(cat_rows, colWidths=[280, 245])
    cat_t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor(navy_hex)),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BOX', (0,0), (-1,-1), 0.8, colors.HexColor(navy_hex)),
    ])
    for idx in range(1, len(cat_rows) - 1):
        bg = colors.HexColor('#F8FAFC') if idx % 2 == 0 else colors.white
        cat_t_style.add('BACKGROUND', (0, idx), (-1, idx), bg)
    cat_t_style.add('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E2E8F0'))
    cat_table.setStyle(cat_t_style)
    story.append(cat_table)
    story.append(PageBreak())
    
    # ── Page 3: JUNE BUDGET AUDIT ──
    story.append(Paragraph("III. June 2026 Project Courier Budget & Expenditure Audit", section_title_style))
    
    june_rows = [[
        Paragraph("Project ID", cell_header_style),
        Paragraph("Client Name", cell_header_style),
        Paragraph("VND ➔ ENG Budget", cell_header_style),
        Paragraph("ENG ➔ CLT Budget", cell_header_style),
        Paragraph("Actual June Dr. (₹)", cell_header_style),
        Paragraph("Variance (₹)", cell_header_style),
        Paragraph("Status", cell_header_style)
    ]]
    for pid, tot in sorted(project_totals_june.items(), key=lambda x: x[1], reverse=True):
        bud_v2e, bud_e2c = courier_budgets.get(pid, (None, None))
        v2e_str = f"₹{bud_v2e:,.2f}" if bud_v2e is not None else "-"
        e2c_str = f"₹{bud_e2c:,.2f}" if bud_e2c is not None else "-"
        
        total_budget = (bud_v2e or 0) + (bud_e2c or 0)
        variance = total_budget - tot
        
        if bud_v2e is None and bud_e2c is None:
            var_para = Paragraph("-", cell_normal_style)
            status_para = Paragraph("No Budget", cell_normal_style)
        elif variance < 0:
            var_para = Paragraph(f"-₹{abs(variance):,.2f}", cell_red_bold_style)
            status_para = Paragraph("🔴 Over Budget", cell_red_bold_style)
        else:
            var_para = Paragraph(f"+₹{variance:,.2f}", cell_green_bold_style)
            status_para = Paragraph("🟢 Under Budget", cell_green_bold_style)
            
        june_rows.append([
            Paragraph(pid, cell_bold_style),
            Paragraph(get_client_name(pid), cell_normal_style),
            Paragraph(v2e_str, cell_normal_style),
            Paragraph(e2c_str, cell_normal_style),
            Paragraph(f"₹{tot:,.2f}", cell_bold_style if tot > 0 else cell_normal_style),
            var_para,
            status_para
        ])
        
    june_table = Table(june_rows, colWidths=[55, 120, 75, 75, 65, 70, 65])
    june_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor(navy_hex)),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 4),
        ('TOPPADDING', (0,0), (-1,0), 4),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BOX', (0,0), (-1,-1), 0.8, colors.HexColor(navy_hex)),
    ]))
    story.append(june_table)
    story.append(PageBreak())
    
    # ── Page 4: JULY BUDGET AUDIT ──
    story.append(Paragraph("IV. July 2026 Project Courier Budget & Expenditure Audit", section_title_style))
    
    july_rows = [[
        Paragraph("Project ID", cell_header_style),
        Paragraph("Client Name", cell_header_style),
        Paragraph("VND ➔ ENG Budget", cell_header_style),
        Paragraph("ENG ➔ CLT Budget", cell_header_style),
        Paragraph("Actual July Dr. (₹)", cell_header_style),
        Paragraph("Variance (₹)", cell_header_style),
        Paragraph("Status", cell_header_style)
    ]]
    for pid, tot in sorted(project_totals_july.items(), key=lambda x: x[1], reverse=True):
        bud_v2e, bud_e2c = courier_budgets.get(pid, (None, None))
        v2e_str = f"₹{bud_v2e:,.2f}" if bud_v2e is not None else "-"
        e2c_str = f"₹{bud_e2c:,.2f}" if bud_e2c is not None else "-"
        
        total_budget = (bud_v2e or 0) + (bud_e2c or 0)
        variance = total_budget - tot
        
        if bud_v2e is None and bud_e2c is None:
            var_para = Paragraph("-", cell_normal_style)
            status_para = Paragraph("No Budget", cell_normal_style)
        elif variance < 0:
            var_para = Paragraph(f"-₹{abs(variance):,.2f}", cell_red_bold_style)
            status_para = Paragraph("🔴 Over Budget", cell_red_bold_style)
        else:
            var_para = Paragraph(f"+₹{variance:,.2f}", cell_green_bold_style)
            status_para = Paragraph("🟢 Under Budget", cell_green_bold_style)
            
        july_rows.append([
            Paragraph(pid, cell_bold_style),
            Paragraph(get_client_name(pid), cell_normal_style),
            Paragraph(v2e_str, cell_normal_style),
            Paragraph(e2c_str, cell_normal_style),
            Paragraph(f"₹{tot:,.2f}", cell_bold_style if tot > 0 else cell_normal_style),
            var_para,
            status_para
        ])
        
    july_table = Table(july_rows, colWidths=[55, 120, 75, 75, 65, 70, 65])
    july_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor(navy_hex)),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 4),
        ('TOPPADDING', (0,0), (-1,0), 4),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BOX', (0,0), (-1,-1), 0.8, colors.HexColor(navy_hex)),
    ]))
    story.append(july_table)
    
    doc.build(story, canvasmaker=NumberedCanvas)
    print("SUCCESS: Consolidated 2026 report PDF generated with Excel Overview.")

if __name__ == '__main__':
    main()
