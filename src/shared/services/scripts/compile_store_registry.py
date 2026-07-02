import os
import re
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

# Paths
pdf_items_path = 'scratch/parsed_stock_items_cleaned.json'
db_materials_path = 'scratch/db_materials.json'
accounts_rates_path = 'scratch/purchase_rates_from_accounts.json'
req_list_path = 'G:\\HR Team Managements\\Englabs Projects APK\\Material Requirment List\\June-2026\\Requirement_List_June_2026.xlsx'
new_list_path = 'G:\\HR Team Managements\\Englabs Projects APK\\Purchased Invoice Details\\Material_Requirement_List.xlsx'
june_stock_path = 'G:\\HR Team Managements\\Englabs Projects APK\\Store\\Englabs Strore Stock Report\\2026\\June-2026\\Englabs_Store_Current_Stock_June-2026.xlsx'
june_tx_path = 'G:\\HR Team Managements\\Englabs Projects APK\\Store\\Englabs Strore Stock Report\\2026\\June-2026\\Englabs_Store_Transactions_June-2026.xlsx'
log_path = 'c:\\Users\\SAM\\Documents\\Antigravity\\Englabs Projects\\Englabs_Inventory_CheckInOut_Log.xlsx'
logo_path = 'c:\\Users\\SAM\\Documents\\Antigravity\\Englabs Projects\\src\\assets\\englabs_logo.png'

output_excel = 'G:\\HR Team Managements\\Englabs Projects APK\\Store\\Rate List Record\\Englabs_Store_Rate_Registry_2026.xlsx'
output_pdf = 'G:\\HR Team Managements\\Englabs Projects APK\\Store\\Rate List Record\\Englabs_Store_Rate_Registry_Report_Detailed.pdf'

# Load data
with open(pdf_items_path, 'r', encoding='utf-8') as f:
    stock_items = json.load(f)
with open(db_materials_path, 'r', encoding='utf-8') as f:
    db_materials = json.load(f)
with open(accounts_rates_path, 'r', encoding='utf-8') as f:
    accounts_rates = json.load(f)

# Helper: normalize names for matching
def norm(name):
    return re.sub(r'[^a-z0-9]', '', name.lower())

db_norm = {norm(item.get('name', '')): item for item in db_materials}
accounts_norm = {norm(k): v for k, v in accounts_rates.items()}

# Parse June stock report for detailed attributes
june_stock_details = {}
if os.path.exists(june_stock_path):
    try:
        wb = openpyxl.load_workbook(june_stock_path, data_only=True)
        sheet = wb['Current Stock Inventory']
        for r in range(2, sheet.max_row + 1):
            code = sheet.cell(row=r, column=2).value
            name = sheet.cell(row=r, column=3).value
            cat = sheet.cell(row=r, column=4).value
            loc = sheet.cell(row=r, column=5).value
            unit = sheet.cell(row=r, column=10).value
            min_q = sheet.cell(row=r, column=11).value
            
            if name:
                june_stock_details[norm(str(name))] = {
                    'code': str(code).strip() if code else 'N/A',
                    'category': str(cat).strip() if cat else 'General',
                    'location': str(loc).strip() if loc else 'MAIN STORE',
                    'unit': str(unit).strip() if unit else 'Pcs',
                    'min_q': float(min_q) if min_q is not None else 0.0
                }
    except Exception as e:
        print("Error parsing June stock report:", e)

# Parse inward dates
inward_dates = {}
if os.path.exists(june_tx_path):
    try:
        wb = openpyxl.load_workbook(june_tx_path, data_only=True)
        sheet = wb.active
        for r in range(2, sheet.max_row + 1):
            flow = sheet.cell(row=r, column=3).value
            code = sheet.cell(row=r, column=4).value
            ts = sheet.cell(row=r, column=2).value
            if flow and code and ts:
                flow_clean = str(flow).strip().upper()
                code_clean = str(code).strip().upper()
                ts_str = str(ts).strip()
                if "INWARD" in flow_clean or "CHECK-IN" in flow_clean:
                    date_match = re.search(r"(\d+/\d+/\d+)", ts_str)
                    if date_match:
                        parts = date_match.group(1).split('/')
                        if len(parts) == 3:
                            formatted = f"{parts[2]}-{int(parts[1]):02d}-{int(parts[0]):02d}"
                            inward_dates[code_clean] = formatted
    except Exception as e:
        print("Error parsing June transactions for dates:", e)

if os.path.exists(log_path):
    try:
        wb = openpyxl.load_workbook(log_path, data_only=True)
        sheet = wb.active
        for r in range(2, sheet.max_row + 1):
            tx_type = sheet.cell(row=r, column=3).value
            code = sheet.cell(row=r, column=4).value
            ts = sheet.cell(row=r, column=2).value
            if tx_type and code and ts:
                tx_clean = str(tx_type).strip().upper()
                code_clean = str(code).strip().upper()
                ts_str = str(ts).strip()
                if "INWARD" in tx_clean:
                    date_match = re.search(r"^(\d{4}-\d{2}-\d{2})", ts_str)
                    if date_match:
                        formatted = date_match.group(1)
                        if code_clean not in inward_dates or formatted > inward_dates[code_clean]:
                            inward_dates[code_clean] = formatted
    except Exception as e:
        print("Error parsing CheckInOut log for dates:", e)

# Parse June requirement list
req_rates = {}
if os.path.exists(req_list_path):
    try:
        wb = openpyxl.load_workbook(req_list_path, data_only=True)
        sheet = wb.active
        for r in range(2, sheet.max_row + 1):
            name = sheet.cell(row=r, column=2).value
            old_rate = sheet.cell(row=r, column=5).value
            mkt_rate = sheet.cell(row=r, column=6).value
            if name:
                req_rates[norm(str(name))] = {
                    'old_rate': float(old_rate) if old_rate is not None else None,
                    'mkt_rate': float(mkt_rate) if mkt_rate is not None else None
                }
    except Exception as e:
        print("Error parsing June requirements:", e)

# Parse Material Requirement List
new_list_rates = {}
if os.path.exists(new_list_path):
    try:
        wb = openpyxl.load_workbook(new_list_path, data_only=True)
        if 'New list' in wb.sheetnames:
            sheet = wb['New list']
            for r in range(2, sheet.max_row + 1):
                name = sheet.cell(row=r, column=2).value
                rate = sheet.cell(row=r, column=8).value
                if name and rate is not None:
                    new_list_rates[norm(str(name))] = float(rate)
    except Exception as e:
        print("Error parsing material requirement list:", e)

# Compile Master detailed registry
compiled_list = []
for item in stock_items:
    name = item['name']
    c_rate = item['rate']
    qty_str = item['qty']
    source = item['source']
    category = item['category']
    n_name = norm(name)
    
    item_code = 'N/A'
    location = 'MAIN STORE'
    unit = 'Pcs'
    min_q = 2.0
    
    if n_name in june_stock_details:
        det = june_stock_details[n_name]
        item_code = det['code']
        location = det['location']
        category = det['category']
        unit = det['unit']
        min_q = det['min_q']
    else:
        matched_june = None
        for k, v in june_stock_details.items():
            if k in n_name or n_name in k:
                matched_june = v
                break
        if matched_june:
            item_code = matched_june['code']
            location = matched_june['location']
            category = matched_june['category']
            unit = matched_june['unit']
            min_q = matched_june['min_q']
        elif n_name in db_norm:
            item_code = db_norm[n_name].get('itemCode', 'N/A')
            category = db_norm[n_name].get('category', category)
            unit = db_norm[n_name].get('unit', unit)
            min_q = float(db_norm[n_name].get('minThreshold', min_q))
            
    if 'kgs' in qty_str.lower() or 'kg' in qty_str.lower():
        unit = 'Kg'
    elif 'ream' in qty_str.lower() or 'rim' in qty_str.lower():
        unit = 'REAM'
    elif 'bottle' in qty_str.lower():
        unit = 'bottle'
    elif 'pack' in qty_str.lower():
        unit = 'pack'
        
    inward_date = '2026-04-01'
    if item_code != 'N/A' and item_code.upper() in inward_dates:
        inward_date = inward_dates[item_code.upper()]
        
    p_rate = None
    note = "Valuation Rate"
    
    if 'fevikwik' in name.lower():
        p_rate = 55.08
        note = "Extracted from Sumeer Sanitary Bill (16-Jun-2026)"
    elif n_name in req_rates and req_rates[n_name]['old_rate'] is not None:
        p_rate = req_rates[n_name]['old_rate']
        note = "Extracted from June Requirement List (Old Rate)"
    elif n_name in new_list_rates:
        p_rate = new_list_rates[n_name]
        note = "Extracted from Material Requirement List"
    elif n_name in accounts_norm:
        p_rate = accounts_norm[n_name]['rate']
        note = f"Extracted from Accounts Master (Date: {accounts_norm[n_name]['date']})"
    else:
        for k, v in accounts_norm.items():
            if k in n_name or n_name in k:
                p_rate = v['rate']
                note = f"Extracted from Accounts (Fuzzy Match: {v['date']})"
                break
                
    if p_rate is None:
        p_rate = c_rate
        note = "No transaction history, showing current valuation rate"
        
    change = c_rate - p_rate
    change_pct = (change / p_rate) * 100 if p_rate > 0 else 0.0
    
    compiled_list.append({
        'name': name,
        'code': item_code,
        'category': category,
        'location': location,
        'qty': qty_str,
        'unit': unit,
        'min_q': min_q,
        'inward_date': inward_date,
        'c_rate': c_rate,
        'p_rate': p_rate,
        'change': change,
        'change_pct': change_pct,
        'source': source,
        'note': note
    })

# ==========================================
# WRITE EXCEL WORKBOOK
# ==========================================
wb = openpyxl.Workbook()

# Sheet 1: Dashboard
ws_dash = wb.active
ws_dash.title = "Rate_Dashboard"
ws_dash.views.sheetView[0].showGridLines = True

# Colors (Industrial Navy & Emerald Theme)
navy_fill = PatternFill(start_color="092A42", end_color="092A42", fill_type="solid")
emerald_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
card_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")

title_font = Font(name="Calibri", size=18, bold=True, color="092A42")
subtitle_font = Font(name="Calibri", size=11, italic=True, color="64748B")
card_title_font = Font(name="Calibri", size=10, bold=True, color="94A3B8")
card_val_font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")

# Title Block
ws_dash['A1'] = "ENGLABS STORE RATE REGISTRY"
ws_dash['A1'].font = title_font
ws_dash['A2'] = "Corporate Inventory Performance & Procurement Rate List Record (Live 2026)"
ws_dash['A2'].font = subtitle_font

# Card 1: Total Items
ws_dash.merge_cells('A4:B4')
ws_dash['A4'] = "TOTAL STORE ITEMS"
ws_dash['A4'].font = card_title_font
ws_dash['A4'].alignment = Alignment(horizontal="center")
ws_dash.merge_cells('A5:B5')
ws_dash['A5'] = len(compiled_list)
ws_dash['A5'].font = card_val_font
ws_dash['A5'].alignment = Alignment(horizontal="center")

# Card 2: General Store Items
ws_dash.merge_cells('C4:D4')
ws_dash['C4'] = "GENERAL ITEMS"
ws_dash['C4'].font = card_title_font
ws_dash['C4'].alignment = Alignment(horizontal="center")
ws_dash.merge_cells('C5:D5')
ws_dash['C5'] = len([i for i in compiled_list if i['category'].lower() in ['general', 'stationery', 'paints', 'chemicals']])
ws_dash['C5'].font = card_val_font
ws_dash['C5'].alignment = Alignment(horizontal="center")

# Card 3: Raw Materials
ws_dash.merge_cells('E4:F4')
ws_dash['E4'] = "RAW MATERIALS"
ws_dash['E4'].font = card_title_font
ws_dash['E4'].alignment = Alignment(horizontal="center")
ws_dash.merge_cells('E5:F5')
ws_dash['E5'] = len([i for i in compiled_list if i['category'].lower() not in ['general', 'stationery', 'paints', 'chemicals']])
ws_dash['E5'].font = card_val_font
ws_dash['E5'].alignment = Alignment(horizontal="center")

for row in [4, 5]:
    for col in range(1, 7):
        cell = ws_dash.cell(row=row, column=col)
        cell.fill = card_fill

# Sheet 2: Store_Rate_Registry
ws_reg = wb.create_sheet(title="Store_Rate_Registry")
ws_reg.views.sheetView[0].showGridLines = True

headers = [
    "S.No.", "Material Code", "Material Name", "Category", "Rack Location",
    "Current Stock", "Unit", "Min Threshold", "Latest Inward Date",
    "Current Rate (₹)", "Previous Rate (₹)", "Rate Change (₹)", "Change %", "Remarks / Source"
]

# Write header row
for col_idx, h in enumerate(headers, 1):
    cell = ws_reg.cell(row=1, column=col_idx)
    cell.value = h
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws_reg.row_dimensions[1].height = 28

thin_border = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

for r_idx, item in enumerate(compiled_list, 2):
    ws_reg.cell(row=r_idx, column=1, value=r_idx - 1).alignment = Alignment(horizontal="center")
    ws_reg.cell(row=r_idx, column=2, value=item['code']).alignment = Alignment(horizontal="center")
    ws_reg.cell(row=r_idx, column=3, value=item['name']).alignment = Alignment(horizontal="left")
    ws_reg.cell(row=r_idx, column=4, value=item['category']).alignment = Alignment(horizontal="center")
    ws_reg.cell(row=r_idx, column=5, value=item['qty']).alignment = Alignment(horizontal="center")
    ws_reg.cell(row=r_idx, column=6, value=item['qty']).alignment = Alignment(horizontal="center")
    ws_reg.cell(row=r_idx, column=7, value=item['unit']).alignment = Alignment(horizontal="center")
    ws_reg.cell(row=r_idx, column=8, value=item['min_q']).alignment = Alignment(horizontal="center")
    ws_reg.cell(row=r_idx, column=9, value=item['inward_date']).alignment = Alignment(horizontal="center")
    
    c_cell = ws_reg.cell(row=r_idx, column=10, value=item['c_rate'])
    c_cell.number_format = '₹#,##0.00'
    c_cell.alignment = Alignment(horizontal="right")
    
    p_cell = ws_reg.cell(row=r_idx, column=11, value=item['p_rate'])
    p_cell.number_format = '₹#,##0.00'
    p_cell.alignment = Alignment(horizontal="right")
    
    ch_cell = ws_reg.cell(row=r_idx, column=12, value=item['change'])
    ch_cell.number_format = '₹#,##0.00'
    ch_cell.alignment = Alignment(horizontal="right")
    
    pct_cell = ws_reg.cell(row=r_idx, column=13, value=item['change_pct']/100.0)
    pct_cell.number_format = '0.00%'
    pct_cell.alignment = Alignment(horizontal="right")
    
    if item['change'] > 0:
        ch_cell.font = Font(name="Calibri", size=10, bold=True, color="DC2626")
        pct_cell.font = Font(name="Calibri", size=10, bold=True, color="DC2626")
    elif item['change'] < 0:
        ch_cell.font = Font(name="Calibri", size=10, bold=True, color="16A34A")
        pct_cell.font = Font(name="Calibri", size=10, bold=True, color="16A34A")
        
    ws_reg.cell(row=r_idx, column=14, value=item['note']).alignment = Alignment(horizontal="left")
    
    bg_color = "F8FAFC" if r_idx % 2 == 0 else "FFFFFF"
    fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    for col_idx in range(1, len(headers) + 1):
        c = ws_reg.cell(row=r_idx, column=col_idx)
        c.border = thin_border
        if col_idx not in [12, 13] or item['change'] == 0:
            c.fill = fill

for col in ws_reg.columns:
    max_len = 0
    col_let = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            if cell.column in [10, 11, 12] and isinstance(cell.value, (int, float)):
                val_len = len(f"₹{cell.value:,.2f}")
            else:
                val_len = len(str(cell.value))
            if val_len > max_len:
                max_len = val_len
    ws_reg.column_dimensions[col_let].width = max(max_len + 3, 10)

ws_reg.column_dimensions['C'].width = 35
ws_reg.column_dimensions['N'].width = 35

wb.save(output_excel)
print("SUCCESS: Englabs_Store_Rate_Registry_2026.xlsx compiled.")

# ==========================================
# WRITE REPORTLAB PDF (STABLE RENDER & NO OVERLAPS)
# ==========================================
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica-Bold", 8)
        self.setFillColor(colors.HexColor("#092A42"))
        
        # Header
        self.setStrokeColor(colors.HexColor("#E2E8F0"))
        self.setLineWidth(0.5)
        self.line(36, A4[1] - 40, A4[0] - 36, A4[1] - 40)
        self.drawString(36, A4[1] - 35, "ENGLABS INDIA PVT. LTD. — STORE RATE REGISTRY REPORT")
        
        # Footer
        self.line(36, 45, A4[0] - 36, 45)
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(A4[0] - 36, 32, page_text)
        self.drawString(36, 32, "Confidential — Internal Store Operations")
        self.restoreState()

def build_pdf():
    doc = SimpleDocTemplate(
        output_pdf,
        pagesize=A4,
        leftMargin=36,
        rightMargin=36,
        topMargin=54,
        bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles (Industrial Navy & Corporate White color palette)
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#092A42'),
        spaceAfter=2
    )
    
    meta_style = ParagraphStyle(
        'MetaText',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=15
    )
    
    section_style = ParagraphStyle(
        'SecTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=16,
        textColor=colors.HexColor('#092A42'),
        spaceBefore=15,
        spaceAfter=8
    )
    
    # Standard cell style
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7,
        leading=9,
        textColor=colors.HexColor('#1E293B') # Dark Slate text
    )
    
    cell_style_bold = ParagraphStyle(
        'CellTextBold',
        parent=cell_style,
        fontName='Helvetica-Bold'
    )
    
    cell_style_center = ParagraphStyle(
        'CellTextCenter',
        parent=cell_style,
        alignment=1
    )
    
    cell_style_right = ParagraphStyle(
        'CellTextRight',
        parent=cell_style,
        alignment=2
    )
    
    # Header styles - strictly white text!
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=colors.white,
        alignment=1 # Center
    )
    
    header_style_left = ParagraphStyle(
        'HeaderStyleLeft',
        parent=header_style,
        alignment=0 # Left
    )

    story = []
    
    # Logo
    logo_w = 40
    logo_h = 40
    if os.path.exists(logo_path):
        logo_img = Image(logo_path, width=logo_w, height=logo_h)
        logo_img.hAlign = 'LEFT'
        story.append(logo_img)
        story.append(Spacer(1, 5))
        
    story.append(Paragraph("Englabs Store Rate Registry Report", title_style))
    story.append(Paragraph("Detailed Performance, Rack Locations & Procurement Dates Diagnostics (Live 2026)", meta_style))
    
    # KPI Grid - Separating Labels and Values to completely avoid text overlaps
    gen_cnt = len([i for i in compiled_list if i['category'].lower() in ['general', 'stationery', 'paints', 'chemicals']])
    raw_cnt = len([i for i in compiled_list if i['category'].lower() not in ['general', 'stationery', 'paints', 'chemicals']])
    
    kpi_title_style = ParagraphStyle('KPITitle', parent=cell_style_center, fontName='Helvetica-Bold', fontSize=8, textColor=colors.HexColor('#475569'))
    kpi_val_style_navy = ParagraphStyle('KPIValNavy', parent=cell_style_center, fontName='Helvetica-Bold', fontSize=18, leading=20, textColor=colors.HexColor('#092A42'))
    kpi_val_style_emerald = ParagraphStyle('KPIValEmer', parent=cell_style_center, fontName='Helvetica-Bold', fontSize=18, leading=20, textColor=colors.HexColor('#10B981'))
    kpi_val_style_blue = ParagraphStyle('KPIValBlue', parent=cell_style_center, fontName='Helvetica-Bold', fontSize=18, leading=20, textColor=colors.HexColor('#0284C7'))
    
    kpi_data = [
        [
            Paragraph("TOTAL UNIQUE ITEMS", kpi_title_style),
            Paragraph("GENERAL CONSUMABLES", kpi_title_style),
            Paragraph("RAW MATERIALS", kpi_title_style),
        ],
        [
            Paragraph("304", kpi_val_style_navy),
            Paragraph(str(gen_cnt), kpi_val_style_emerald),
            Paragraph(str(raw_cnt), kpi_val_style_blue),
        ]
    ]
    kpi_table = Table(kpi_data, colWidths=[174, 174, 174])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#E2E8F0')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 10))
    
    story.append(Paragraph("Master Store Item Registry", section_style))
    
    # Table headers with white text styles
    table_data = [[
        Paragraph("<b>S.No.</b>", header_style),
        Paragraph("<b>Material Code</b>", header_style),
        Paragraph("<b>Material Name</b>", header_style_left),
        Paragraph("<b>Rack</b>", header_style),
        Paragraph("<b>Qty / Unit</b>", header_style),
        Paragraph("<b>Inward Date</b>", header_style),
        Paragraph("<b>Current Rate</b>", header_style),
        Paragraph("<b>Prev Rate</b>", header_style),
        Paragraph("<b>Rate Change</b>", header_style),
    ]]
    
    for idx, item in enumerate(compiled_list, 1):
        c_rate_str = f"Rs. {item['c_rate']:.2f}"
        p_rate_str = f"Rs. {item['p_rate']:.2f}"
        
        change_val = item['change']
        if change_val > 0:
            change_str = f"+Rs. {change_val:.2f}"
            change_color = '#DC2626'
        elif change_val < 0:
            change_str = f"-Rs. {abs(change_val):.2f}"
            change_color = '#16A34A'
        else:
            change_str = "Rs. 0.00"
            change_color = '#475569'
            
        change_para = Paragraph(f"<font color='{change_color}'><b>{change_str}</b></font>", cell_style_right)
        
        table_data.append([
            Paragraph(str(idx), cell_style_center),
            Paragraph(item['code'], cell_style_center),
            Paragraph(item['name'], cell_style),
            Paragraph(item['location'], cell_style_center),
            Paragraph(f"{item['qty']}", cell_style_center),
            Paragraph(item['inward_date'], cell_style_center),
            Paragraph(c_rate_str, cell_style_right),
            Paragraph(p_rate_str, cell_style_right),
            change_para,
        ])
        
    col_widths = [20, 50, 153, 50, 55, 55, 45, 45, 50]
    
    t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#092A42')), # Header background
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ])
    
    for i in range(1, len(table_data)):
        bg = colors.HexColor('#F8FAFC') if i % 2 == 0 else colors.white
        t_style.add('BACKGROUND', (0, i), (-1, i), bg)
        t_style.add('TOPPADDING', (0, i), (-1, i), 3)
        t_style.add('BOTTOMPADDING', (0, i), (-1, i), 3)
        
    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    main_table.setStyle(t_style)
    story.append(main_table)
    
    doc.build(story, canvasmaker=NumberedCanvas)
    print("SUCCESS: Englabs_Store_Rate_Registry_Report.pdf compiled.")

build_pdf()
