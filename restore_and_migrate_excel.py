import os
import shutil
import openpyxl
from datetime import datetime

original_file = r"G:\HR Team Managements\Englabs Projects APK\Projects\Current Projects  DETAISL.xlsx"
backup_file = r"G:\HR Team Managements\Englabs Projects APK\Projects\Current Projects  DETAISL_backup.xlsx"
new_sync_file = r"G:\HR Team Managements\Englabs Projects APK\Accounts\Englabs_Projects_Outlook_Sync.xlsx"

projects_to_add = [
    {'Project ID': 'EQ5487', 'Client Name': 'EndureAir', 'Status': 'Pending', 'Date Added': datetime.now().strftime('%Y-%m-%d')},
    {'Project ID': 'EQ5488', 'Client Name': 'RND Mechanical', 'Status': 'Pending', 'Date Added': datetime.now().strftime('%Y-%m-%d')},
    {'Project ID': 'EQ5489', 'Client Name': 'Aebocode', 'Status': 'Pending', 'Date Added': datetime.now().strftime('%Y-%m-%d')},
    {'Project ID': 'EQ5490', 'Client Name': 'Orient Electric 1', 'Status': 'Pending', 'Date Added': datetime.now().strftime('%Y-%m-%d')},
    {'Project ID': 'EQ5491', 'Client Name': 'Orient Electric 2', 'Status': 'Pending', 'Date Added': datetime.now().strftime('%Y-%m-%d')}
]

try:
    # 1. Restore Original File
    if os.path.exists(backup_file):
        shutil.copy2(backup_file, original_file)
        print(f"Success: Restored '{original_file}' from backup.")
    else:
        print("Error: Backup file not found. Could not restore.")

    # 2. Ensure Accounts directory exists
    os.makedirs(os.path.dirname(new_sync_file), exist_ok=True)

    # 3. Create New Workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Outlook Sync Projects"

    # Add Headers
    headers = ["Sr No", "Project ID", "Client Name", "Status", "Date Added"]
    for col_idx, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_idx).value = header

    # Add Projects
    for i, proj in enumerate(projects_to_add):
        row_idx = i + 2
        sheet.cell(row=row_idx, column=1).value = i + 1
        sheet.cell(row=row_idx, column=2).value = proj['Project ID']
        sheet.cell(row=row_idx, column=3).value = proj['Client Name']
        sheet.cell(row=row_idx, column=4).value = proj['Status']
        sheet.cell(row=row_idx, column=5).value = proj['Date Added']

    wb.save(new_sync_file)
    print(f"Success: Created new workbook at '{new_sync_file}'")
    print(f"Success: Inserted {len(projects_to_add)} records into the new workbook.")

except Exception as e:
    print(f"Error executing rollback and migration: {e}")
