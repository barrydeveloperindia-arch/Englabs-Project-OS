import openpyxl
import os

filePath = r"G:\HR Team Managements\Englabs Projects APK\Projects\23-06-2026_Project_Tracker.xlsx"

def find_project_row(sheet, col_idx, pid):
    for r in range(1, sheet.max_row + 1):
        val = sheet.cell(row=r, column=col_idx).value
        if val and str(val).strip() == str(pid).strip():
            return r
    return None

def main():
    if not os.path.exists(filePath):
        print(f"Error: Excel tracker not found at {filePath}")
        return
        
    print(f"Loading workbook: {filePath}...")
    wb = openpyxl.load_workbook(filePath, data_only=False)
    
    # 1. Update Current Projects sheet dynamically for C5254
    sheet1 = wb['Current Projects']
    c5254_row = find_project_row(sheet1, 8, "C5254")
    if c5254_row:
        print(f"Updating C5254 in 'Current Projects' sheet on dynamic row {c5254_row}...")
        # Col 40 (AN) = Vendor Details
        sheet1.cell(row=c5254_row, column=40, value="EXCEL RAPIDTECH PRIVATE LIMITED")
        # Col 46 (AT) = PO Vender Sent
        sheet1.cell(row=c5254_row, column=46, value="P2100-EL12-PO-16953")
        # Col 48 (AV) = WITHOUT Prices
        sheet1.cell(row=c5254_row, column=48, value=5699.60)
    else:
        print("Warning: C5254 row not found in 'Current Projects'")
        
    # 2. Update Vendor PO Release sheet dynamically for C5254
    # Note: sheet name might be 'Vendor PO Release' or 'Vendor PO Relese' (handle spelling safely)
    sname = 'Vendor PO Release' if 'Vendor PO Release' in wb.sheetnames else 'Vendor PO Relese'
    sheet2 = wb[sname]
    # In Vendor PO Release sheet, Quote ID / Project ID is in Column 4 (D)
    po_row = find_project_row(sheet2, 4, "C5254")
    if po_row:
        print(f"Updating C5254 in '{sname}' sheet on dynamic row {po_row}...")
        # Col 8 (H) = Vendor Name
        sheet2.cell(row=po_row, column=8, value="EXCEL RAPIDTECH PRIVATE LIMITED")
        # Col 9 (I) = Vendor Purchased Prices
        sheet2.cell(row=po_row, column=9, value=5699.60)
    else:
        print(f"Warning: C5254 row not found in '{sname}'")
        
    # Save the workbook
    print("Saving workbook...")
    wb.save(filePath)
    print("Workbook successfully updated dynamically and saved!")

if __name__ == "__main__":
    main()
