import os
import json
import openpyxl
import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Define G Drive target base path
G_DRIVE_BASE = r"G:\HR Team Managements\Englabs Projects APK\Projects"
DATA_DIR = "./data"

def create_styled_excel(file_path, sheet_name, headers, rows):
    """Creates a beautifully styled Excel workbook using openpyxl."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Fonts and Colors
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10, color="000000")
    title_font = Font(name="Segoe UI", size=14, bold=True, color="0B3C5D")
    subtitle_font = Font(name="Segoe UI", size=8, italic=True, color="7F8C8D")
    
    header_fill = PatternFill(start_color="0E4368", end_color="0E4368", fill_type="solid")
    even_row_fill = PatternFill(start_color="F2F7FA", end_color="F2F7FA", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    # Title Block
    max_col = len(headers)
    max_col_letter = openpyxl.utils.get_column_letter(max_col)
    
    ws.merge_cells(f"A1:{max_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = f"Englabs Pvt Ltd - {sheet_name}"
    title_cell.font = title_font
    ws.row_dimensions[1].height = 25
    
    ws.merge_cells(f"A2:{max_col_letter}2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Generated automatically on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    sub_cell.font = subtitle_font
    ws.row_dimensions[2].height = 18
    
    # Blank row
    ws.row_dimensions[3].height = 12
    
    # Headers (Row 4)
    ws.row_dimensions[4].height = 28
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    # Rows (starting at Row 5)
    for r_idx, row_data in enumerate(rows, 5):
        ws.row_dimensions[r_idx].height = 22
        use_fill = (r_idx % 2 == 0)
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            
            if use_fill:
                cell.fill = even_row_fill
                
            # Alignments & Formats
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                # Detect columns to apply currency formatting (containing Price, Value, Cost, Margin, Porter, Budget)
                col_header = headers[c_idx - 1].upper()
                if "INR" in col_header or "PRICE" in col_header or "VALUE" in col_header or "COST" in col_header or "MARGIN" in col_header or "PAYMENT" in col_header or "BUDGET" in col_header or "VALUATION" in col_header:
                    cell.number_format = '₹#,##0.00'
                elif "%" in col_header or "PERCENT" in col_header:
                    cell.number_format = '0"%"'
            elif c_idx in [1, 2] or (c_idx == max_col):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.row > 3 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Save file
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb.save(file_path)
    print(f"Created styled spreadsheet at: {file_path}")

def main():
    if not os.path.exists(DATA_DIR):
        print(f"Error: data directory {DATA_DIR} does not exist.")
        return
        
    print("Reading project JSON databases...")
    project_files = [f for f in os.listdir(DATA_DIR) if f.startswith('C') and f.endswith('.json')]
    
    po_rows = []
    invoice_rows = []
    budgets_rows = []
    tracker_rows = []
    standup_rows = []
    lookup_rows = []
    directory_rows = []
    dashboard_rows = []
    
    for filename in project_files:
        file_path = os.path.join(DATA_DIR, filename)
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                p = json.load(f)
                
            project_id = p.get("projectId", "")
            client = p.get("client", "")
            planning = p.get("planning", {})
            client_po = planning.get("poNumber", "")
            base_value = planning.get("value", 0)
            budget_val = planning.get("budget", 0)
            start_date = planning.get("startDate", "—")
            delivery_terms = planning.get("deliveryTerms", "—")
            po_confirmed = "Yes" if planning.get("poConfirmed") else "No"
            
            metrics = p.get("metrics", {})
            materials = metrics.get("materialConsumption") or "—"
            qty = metrics.get("totalComponents", 0)
            
            prod = p.get("production", {})
            current_stage = prod.get("currentStage", "—")
            
            # Check completed
            stages = prod.get("stages", [])
            completed_stages = [s.get("name") for s in stages if s.get("status") == "Completed"]
            pending_stages = [s.get("name") for s in stages if s.get("status") != "Completed"]
            total_stages = len(stages)
            pct_complete = int(round((len(completed_stages) / total_stages) * 100)) if total_stages > 0 else 0
            overall_status = "Completed" if pct_complete == 100 else "Active"
            
            fin = p.get("financials", {})
            v_name = fin.get("vendorName", "—")
            v_loc = fin.get("vendorLocation", "—")
            
            standup = p.get("dailyStandup", {})
            notes = standup.get("discussingNotes", "—")
            prep_date = standup.get("preparingPartsDate", "—")
            r_from = standup.get("routeFrom", "—")
            r_to = standup.get("routeTo", "—")
            porter_pay = standup.get("porterPayments", 0)
            blocker = standup.get("inputsRequired", "None")
            lead = stages[0].get("lead", "—") if stages else "—"
            
            # PO Release data extraction
            po = p.get("poRelease", {})
            sale = po.get("customerSalePrice") or base_value or 0
            cost = po.get("vendorCost", 0)
            margin = sale - cost
            margin_pct = int(round((margin / sale) * 100)) if sale > 0 else 0
            
            # Invoice Release data extraction
            inv = p.get("invoiceRelease", {})
            gst_rate = inv.get("gstRate", 18)
            gst_val = inv.get("gstValue") or (base_value * (gst_rate / 100.0))
            total_val = inv.get("totalInvoiceAmount") or (base_value + gst_val)
            inv_no = inv.get("invoiceNumber", "")
            payment_terms = inv.get("paymentTerms", "—")
            payment_status = inv.get("paymentStatus", "Pending")
            
            # 1. Dashboard Rows (General overview)
            dashboard_rows.append([
                len(dashboard_rows) + 1,
                project_id,
                client,
                base_value,
                budget_val,
                overall_status,
                current_stage,
                pct_complete
            ])
            
            # 2. Directory Rows
            directory_rows.append([
                len(directory_rows) + 1,
                project_id,
                client,
                base_value,
                budget_val,
                materials,
                qty,
                start_date
            ])
            
            # 3. Project Search Rows (Full detail lookup format)
            lookup_rows.append([
                len(lookup_rows) + 1,
                project_id,
                client,
                client_po or "—",
                base_value,
                budget_val,
                v_name,
                po.get("poVendorSent", "—") if po else "—",
                inv_no or "—",
                payment_status
            ])
            
            # 4. Project Budgets Rows
            proj_margin = base_value - budget_val
            proj_margin_pct = int(round((proj_margin / base_value) * 100)) if base_value > 0 else 0
            budgets_rows.append([
                len(budgets_rows) + 1,
                project_id,
                client,
                base_value,
                budget_val,
                proj_margin,
                proj_margin_pct,
                client_po or "—",
                po_confirmed,
                start_date,
                delivery_terms
            ])
            
            # 5. Projects Tracker Rows
            tracker_rows.append([
                len(tracker_rows) + 1,
                project_id,
                client,
                base_value,
                budget_val,
                materials,
                qty,
                current_stage,
                overall_status,
                v_name,
                v_loc,
                client_po or "—",
                start_date
            ])
            
            # 6. Daily Standup Rows
            standup_rows.append([
                len(standup_rows) + 1,
                project_id,
                client,
                notes,
                prep_date,
                r_from,
                r_to,
                porter_pay,
                blocker,
                lead
            ])
            
            # 7. PO Release Rows
            if po and po.get("vendorName"):
                po_rows.append([
                    len(po_rows) + 1,
                    project_id,
                    client,
                    client_po or "—",
                    sale,
                    po.get("vendorName", ""),
                    po.get("vendorLocation", "—"),
                    po.get("poVendorSent", "—"),
                    cost,
                    margin,
                    margin_pct,
                    po.get("releaseDate", "—")
                ])
                
            # 8. Invoice Release Rows
            if inv and inv.get("invoiceNumber"):
                invoice_rows.append([
                    len(invoice_rows) + 1,
                    project_id,
                    client,
                    inv_no,
                    base_value,
                    gst_rate,
                    gst_val,
                    total_val,
                    payment_terms,
                    payment_status
                ])
                
        except Exception as ex:
            print(f"Error parsing {filename}: {ex}")
            
    # Setup G Drive subfolders for each dashboard category
    folders = {
        "Projects Dashboard": os.path.join(G_DRIVE_BASE, "Projects Dashboard"),
        "Project Directory": os.path.join(G_DRIVE_BASE, "Project Directory"),
        "Project Search": os.path.join(G_DRIVE_BASE, "Project Search"),
        "Project Budgets": os.path.join(G_DRIVE_BASE, "Project Budgets"),
        "Projects Tracker": os.path.join(G_DRIVE_BASE, "Projects Tracker"),
        "Daily Standup": os.path.join(G_DRIVE_BASE, "Daily Standup"),
        "PO Release": os.path.join(G_DRIVE_BASE, "PO Release"),
        "Vendor PO Release": os.path.join(G_DRIVE_BASE, "Vendor PO Release"),
        "Invoice Release": os.path.join(G_DRIVE_BASE, "Invoice Release")
    }
    
    # Create the folders if they don't exist
    for dir_path in folders.values():
        os.makedirs(dir_path, exist_ok=True)
        print(f"Initialized folder: {dir_path}")
        
    # --- WRITE SPREADSHEETS ---
    
    # 1. Projects Dashboard
    create_styled_excel(
        os.path.join(folders["Projects Dashboard"], "Projects_Dashboard_Summary.xlsx"),
        "Projects Dashboard Summary",
        ["Sr. No.", "Project ID", "Client", "Valuation (INR)", "Budget Allocation (INR)", "Status", "Current Stage", "Completion Pct (%)"],
        dashboard_rows
    )
    
    # 2. Project Directory
    create_styled_excel(
        os.path.join(folders["Project Directory"], "Project_Directory_Roster.xlsx"),
        "Project Directory Roster",
        ["Sr. No.", "Project ID", "Client", "Valuation (INR)", "Budget Allocation (INR)", "Materials", "Components Qty", "Start Date"],
        directory_rows
    )
    
    # 3. Project Search
    create_styled_excel(
        os.path.join(folders["Project Search"], "Project_Lookup_Report.xlsx"),
        "Project Lookup & Mapping Report",
        ["Sr. No.", "Project ID", "Client", "Client PO Number", "Valuation (INR)", "Budget (INR)", "Vendor Name", "Released Vendor PO", "Invoice Number", "Payment Status"],
        lookup_rows
    )
    
    # 4. Project Budgets
    create_styled_excel(
        os.path.join(folders["Project Budgets"], "Project_Budgets_Audit.xlsx"),
        "Project Budgets Audit Register",
        ["Sr. No.", "Project ID", "Client", "Valuation (INR)", "Budget Allocation (INR)", "Project Margin (INR)", "Margin Pct (%)", "PO Number", "PO Confirmed", "Start Date", "Delivery Terms"],
        budgets_rows
    )
    
    # 5. Projects Tracker
    create_styled_excel(
        os.path.join(folders["Projects Tracker"], "Projects_Tracker_Ledger.xlsx"),
        "Projects Tracker Ledger",
        ["Sr. No.", "Project ID", "Client", "Valuation (INR)", "Budget (INR)", "Materials", "Components Qty", "Current Stage", "Overall Status", "Vendor Name", "Vendor Location", "Client PO Number", "Start Date"],
        tracker_rows
    )
    
    # 6. Daily Standup
    create_styled_excel(
        os.path.join(folders["Daily Standup"], "Daily_Standup_Logs.xlsx"),
        "Daily Standup Logs Register",
        ["Sr. No.", "Project ID", "Client", "Standup Discussion Notes", "Preparing Date", "Route From", "Route To", "Porter Payment (INR)", "Inputs/Blockers Required", "Coordinating Lead"],
        standup_rows
    )
    
    # 7. PO Release / Vendor PO Release
    po_headers = [
        "Sr. No.", "Project ID", "Client", "Client PO Number", 
        "Customer Sale Price (INR)", "Vendor Name", "Vendor Location", 
        "Released Vendor PO", "Vendor Cost Price (INR)", "Gross Margin (INR)", 
        "Margin (%)", "Release Date"
    ]
    if po_rows:
        create_styled_excel(
            os.path.join(folders["PO Release"], "PO_Release.xlsx"),
            "Vendor PO Release Ledger",
            po_headers,
            po_rows
        )
        create_styled_excel(
            os.path.join(folders["Vendor PO Release"], "Vendor_PO_Release.xlsx"),
            "Vendor PO Release Ledger",
            po_headers,
            po_rows
        )
        
    # 8. Invoice Release
    inv_headers = [
        "Sr. No.", "Project ID", "Client", "Invoice Number", 
        "Base Value (INR)", "GST Rate (%)", "GST Value (INR)", 
        "Total Invoice Amount (INR)", "Payment Terms", "Payment Status"
    ]
    if invoice_rows:
        create_styled_excel(
            os.path.join(folders["Invoice Release"], "Invoice_Release.xlsx"),
            "Invoice Release Ledger",
            inv_headers,
            invoice_rows
        )

    print("\nAll Projects sidebar folders and Excel reports successfully updated!")

if __name__ == "__main__":
    main()
