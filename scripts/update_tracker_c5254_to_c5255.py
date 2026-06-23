import openpyxl
import os
import datetime

filePath = r"G:\HR Team Managements\Englabs Projects APK\Projects\23-06-2026_Project_Tracker.xlsx"

def find_project_row(sheet, col_idx, pid):
    for r in range(1, sheet.max_row + 1):
        val = sheet.cell(row=r, column=col_idx).value
        if val and str(val).strip() == str(pid).strip():
            return r
    return None

def find_first_empty_row(sheet, col_idx, start_row=4):
    r = start_row
    while True:
        val = sheet.cell(row=r, column=col_idx).value
        if val is None or str(val).strip() == "":
            return r
        r += 1

def main():
    if not os.path.exists(filePath):
        print(f"Error: {filePath} not found.")
        return
        
    print("Loading workbook...")
    wb = openpyxl.load_workbook(filePath)
    
    # 1. Revert Current Projects for C5254
    sheet1 = wb['Current Projects']
    c5254_row = find_project_row(sheet1, 8, "C5254")
    if c5254_row:
        print(f"Reverting C5254 in 'Current Projects' on dynamic row {c5254_row}...")
        sheet1.cell(row=c5254_row, column=40, value="Excel Repid TECH")
        sheet1.cell(row=c5254_row, column=46, value="N/A")
        sheet1.cell(row=c5254_row, column=48, value=0.00)
    else:
        print("Warning: C5254 row not found in 'Current Projects'")
    
    # 2. Add C5255 to Current Projects dynamically
    c5255_row = find_project_row(sheet1, 8, "C5255")
    if not c5255_row:
        c5255_row = find_first_empty_row(sheet1, 8)
        print(f"Appending C5255 in 'Current Projects' to empty row {c5255_row}...")
    else:
        print(f"Updating existing C5255 in 'Current Projects' on row {c5255_row}...")
        
    sheet1.cell(row=c5255_row, column=1, value=c5255_row - 4) # Dynamic Sr. No. based on offset
    sheet1.cell(row=c5255_row, column=2, value="BAJAJ") # Customer Details
    sheet1.cell(row=c5255_row, column=4, value="MUMBAI") # Location
    sheet1.cell(row=c5255_row, column=5, value="Shubham") # Lead
    sheet1.cell(row=c5255_row, column=7, value="23-06-2026") # Date
    sheet1.cell(row=c5255_row, column=8, value="C5255") # Project ID
    sheet1.cell(row=c5255_row, column=9, value=9010) # Customer Value
    sheet1.cell(row=c5255_row, column=10, value="No")
    sheet1.cell(row=c5255_row, column=11, value="23-06-2026")
    sheet1.cell(row=c5255_row, column=12, value="Shubham")
    sheet1.cell(row=c5255_row, column=13, value="YES")
    sheet1.cell(row=c5255_row, column=24, value="handle_gasket-C5255") # Materials
    sheet1.cell(row=c5255_row, column=25, value="NOS")
    sheet1.cell(row=c5255_row, column=26, value=20) # Qty
    sheet1.cell(row=c5255_row, column=27, value=9010)
    sheet1.cell(row=c5255_row, column=28, value="30 days")
    sheet1.cell(row=c5255_row, column=30, value=5699.60)
    sheet1.cell(row=c5255_row, column=39, value="C5255")
    sheet1.cell(row=c5255_row, column=40, value="EXCEL RAPIDTECH PRIVATE LIMITED") # Vendor Name
    sheet1.cell(row=c5255_row, column=41, value="Delhi") # Location
    sheet1.cell(row=c5255_row, column=46, value="P2100-EL12-PO-16953") # PO Vendor Sent
    sheet1.cell(row=c5255_row, column=47, value="handle_gasket-C5255")
    sheet1.cell(row=c5255_row, column=48, value=5699.60) # Vendor Cost Pre-GST
    sheet1.cell(row=c5255_row, column=49, value="NOS")
    sheet1.cell(row=c5255_row, column=50, value=20)
    
    # 3. Update Vendor PO Release dynamically (relabel C5244 to C5255 if exists)
    sheet2 = wb['Vendor PO Release']
    # Check if we have C5254 in Column 4 (D)
    po_row = find_project_row(sheet2, 4, "C5254")
    if po_row:
        print(f"Updating Vendor PO Release on row {po_row} from C5254 to C5255...")
        sheet2.cell(row=po_row, column=4, value="C5255")
    else:
        # Check if C5255 already exists
        po_row_5255 = find_project_row(sheet2, 4, "C5255")
        if not po_row_5255:
            po_row_5255 = find_first_empty_row(sheet2, 4)
            print(f"Appending C5255 to Vendor PO Release empty row {po_row_5255}...")
            sheet2.cell(row=po_row_5255, column=4, value="C5255")
            
    # 4. Add C5255 to Daily Standup dynamically
    sheet3 = wb['Daily Standup']
    standup_row = find_project_row(sheet3, 2, "C5255")
    if not standup_row:
        standup_row = find_first_empty_row(sheet3, 2)
        print(f"Writing C5255 in 'Daily Standup' to empty row {standup_row}...")
    else:
        print(f"Updating existing C5255 in 'Daily Standup' on row {standup_row}...")
        
    sheet3.cell(row=standup_row, column=1, value=standup_row - 3) # Dynamic Sr. No
    sheet3.cell(row=standup_row, column=2, value="C5255") # Quote ID
    sheet3.cell(row=standup_row, column=3, value=9010) # Project Value
    sheet3.cell(row=standup_row, column=4, value="BAJAJ") # Client
    sheet3.cell(row=standup_row, column=5, value="handle_gasket-C5255") # Finish
    sheet3.cell(row=standup_row, column=6, value="YES / 20") # Qty
    sheet3.cell(row=standup_row, column=7, value="23-06-2026")
    sheet3.cell(row=standup_row, column=8, value="NO") # Blocker
    sheet3.cell(row=standup_row, column=9, value="Shubham") # Staff
    sheet3.cell(row=standup_row, column=10, value="All 20 parts have been completed and are ready for dispatch.")
    sheet3.cell(row=standup_row, column=11, value="2026-06-23")
    sheet3.cell(row=standup_row, column=12, value="C5255") # Current Project
    sheet3.cell(row=standup_row, column=13, value="Shubham") # Porter Staff
    sheet3.cell(row=standup_row, column=15, value="Delhi") # From
    
    print("Saving workbook...")
    wb.save(filePath)
    print("Successfully updated Excel spreadsheet dynamically!")

if __name__ == "__main__":
    main()
