import os
import re
import json
import shutil
import datetime
import openpyxl
import pandas as pd
from PIL import Image
import pytesseract
import fitz  # PyMuPDF

# Paths configuration
WORKSPACE_DIR = r"c:\Users\SAM\Documents\Antigravity\Englabs Projects"
ENQUIRY_FILE = r"C:\Users\SAM\OneDrive - Englabs India Pvt Ltd\2025 ENQUIRIES\ENQUIRIES REGISTER_SA UPDATED - 01-06.xlsx"
ONEDRIVE_DIR = r"C:\Users\SAM\OneDrive - Englabs India Pvt Ltd\2025 ENQUIRIES"
SITE_CASH_FILE = r"G:\Englabs_HR Team Managements\Site Cash Details\Englabs Site Cash.xlsx"
BACKUP_DIR = r"G:\Englabs_HR Team Managements\Site Cash Details\Backups"
HISTORY_FILE = r"G:\Englabs_HR Team Managements\Site Cash Details\update_history.json"
CACHE_FILE = os.path.join(WORKSPACE_DIR, "data", "project_directory_cache.json")
OUTLOOK_DB_FILE = os.path.join(WORKSPACE_DIR, "data", "outlook_processed_db.json")

# Ensure Tesseract path is set
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

def extract_project_digits(name):
    """Extract project digits (4-5 digits), prioritizing EQ/C prefixes and avoiding date years."""
    match = re.search(r'(?:EQ|C)\s*(\d{4,5})', name, re.IGNORECASE)
    if match:
        return match.group(1)
        
    # Fallback to any 4-5 digit number that is not a standard date year
    candidates = re.findall(r'\b\d{4,5}\b', name)
    for c in candidates:
        if c not in ("2024", "2025", "2026"):
            return c
    return None

def build_directory_cache():
    """Crawl OneDrive folders, find client subdirectories, and parse any quote Excel files."""
    print("Building project directory cache from OneDrive...")
    cache = {}
    
    # We scan ENQUIRIES 2025 and ENQUIRIES 2026
    subfolders_to_scan = ["ENQUIRIES 2025", "ENQUIRIES 2026"]
    
    for sub in subfolders_to_scan:
        sub_path = os.path.join(ONEDRIVE_DIR, sub)
        if not os.path.exists(sub_path):
            continue
            
        # List all clients (first level subdirectories under ENQUIRIES 2025/2026)
        try:
            clients = [d for d in os.listdir(sub_path) if os.path.isdir(os.path.join(sub_path, d))]
        except Exception as e:
            print(f"Error listing clients in {sub}: {e}")
            continue
            
        for client in clients:
            client_path = os.path.join(sub_path, client)
            # Walk client folder to find all directories and files
            for root, dirs, files in os.walk(client_path):
                # Search for specific enquiry folders containing digits
                for d in dirs:
                    digits = extract_project_digits(d)
                    if digits:
                        folder_full_path = os.path.join(root, d)
                        
                        # Add baseline entry
                        if digits not in cache:
                            cache[digits] = {
                                "digits": digits,
                                "client_folder_name": client,
                                "project_folder_path": folder_full_path,
                                "year": "2025" if "2025" in sub else "2026",
                                "quote_file": None,
                                "address": "N/A",
                                "city": "N/A",
                                "state": "N/A",
                                "contact_person": "N/A",
                                "mobile_number": "N/A"
                            }
                
                # Check for quotation excel files inside client folders
                for f in files:
                    if (f.endswith('.xlsx') or f.endswith('.xls')) and not f.startswith('~'):
                        digits = extract_project_digits(f)
                        if digits:
                            file_full_path = os.path.join(root, f)
                            
                            # Ensure we have a baseline entry
                            if digits not in cache:
                                cache[digits] = {
                                    "digits": digits,
                                    "client_folder_name": client,
                                    "project_folder_path": root,
                                    "year": "2025" if "2025" in sub else "2026",
                                    "quote_file": None,
                                    "address": "N/A",
                                    "city": "N/A",
                                    "state": "N/A",
                                    "contact_person": "N/A",
                                    "mobile_number": "N/A"
                                }
                            
                            cache[digits]["quote_file"] = file_full_path
                            # Parse details from quote sheet
                            quote_details = parse_quote_details(file_full_path)
                            if quote_details:
                                cache[digits].update(quote_details)
                                
    # Write cache
    os.makedirs(os.path.dirname(CACHE_FILE), exist_ok=True)
    with open(CACHE_FILE, 'w', encoding='utf-8') as f:
        json.dump(cache, f, indent=4, ensure_ascii=False)
        
    print(f"Directory cache saved to {CACHE_FILE}. Total records: {len(cache)}")
    return cache

def parse_quote_details(file_path):
    """Attempt to parse client contact info from quote Excel sheets."""
    try:
        xl = pd.ExcelFile(file_path)
        sheet_name = None
        for s in xl.sheet_names:
            if 'QUOTE' in s.upper() or 'QUOTATION' in s.upper() or 'INVOICE' in s.upper():
                sheet_name = s
                break
        if not sheet_name:
            sheet_name = xl.sheet_names[0]
            
        df = xl.parse(sheet_name, header=None)
        
        details = {
            "address": "N/A",
            "city": "N/A",
            "state": "N/A",
            "contact_person": "N/A",
            "mobile_number": "N/A"
        }
        
        client_row = None
        atten_row = None
        mob_row = None
        
        # Scan cells to find key labels
        for r_idx in range(len(df)):
            for c_idx in range(min(5, len(df.columns))):
                val = str(df.iloc[r_idx, c_idx]).strip()
                if not val or val == 'nan':
                    continue
                
                # Look for Client
                if re.match(r'^client\b', val, re.IGNORECASE) and client_row is None:
                    client_row = r_idx
                # Look for Attn or Atten
                if re.match(r'^attn|^atten', val, re.IGNORECASE) and atten_row is None:
                    atten_row = r_idx
                # Look for Mobile or MOB
                if re.match(r'^mob|^mobile|^phone', val, re.IGNORECASE) and mob_row is None:
                    mob_row = r_idx
                    
        # Extract client name & address
        if client_row is not None:
            # Client address lines typically sit in column 1 below client name
            address_lines = []
            max_addr_row = atten_row if atten_row is not None else client_row + 6
            for r in range(client_row + 1, min(max_addr_row, len(df))):
                addr_val = str(df.iloc[r, 1]).strip()
                if addr_val and addr_val != 'nan':
                    address_lines.append(addr_val)
            if address_lines:
                details["address"] = ", ".join(address_lines)
                # Parse city, state from last address line
                last_line = address_lines[-1]
                # Pattern: City Pincode or City State
                city_match = re.search(r'([A-Za-z\s]+)\s+(\d{6})', last_line)
                if city_match:
                    details["city"] = city_match.group(1).strip()
                    details["state"] = "Punjab/Haryana"  # Default region for Englabs
                else:
                    details["city"] = last_line.split(',')[0].strip()
                    
        # Extract contact person
        if atten_row is not None:
            val = str(df.iloc[atten_row, 1]).strip()
            if val and val != 'nan':
                details["contact_person"] = val
                
        # Extract mobile number
        if mob_row is not None:
            val = str(df.iloc[mob_row, 1]).strip()
            if val and val != 'nan':
                # clean non-digits
                details["mobile_number"] = re.sub(r'\D', '', val)
                
        return details
    except Exception as e:
        # Silently fail for parsing errors, fallback to default NA
        return None

def load_enquiries():
    """Load enquiries from Master List sheet of main database."""
    if not os.path.exists(ENQUIRY_FILE):
        print(f"Enquiry file not found: {ENQUIRY_FILE}")
        return {}
        
    try:
        df = pd.read_excel(ENQUIRY_FILE, sheet_name='MASTER LIST')
        enquiries = {}
        # Ensure string type for EQ NO
        df['EQ NO_str'] = df['EQ NO'].astype(str)
        
        for idx, row in df.iterrows():
            eq_no = str(row.get('EQ NO', ''))
            if not eq_no or eq_no == 'nan':
                continue
                
            # Extract digits from EQ NO
            match = re.search(r'\d+', eq_no)
            if not match:
                continue
            digits = match.group(0)
            
            # Map columns
            from_col = row.get(' FROM', '') # note leading space
            desc = row.get('DESCRIPTION', '')
            date_val = row.get('DATE', '')
            status = row.get('STATUS', 'Pending')
            value = row.get('VALUE', 0)
            
            # Clean values
            from_str = str(from_col).strip() if pd.notna(from_col) else ""
            desc_str = str(desc).strip() if pd.notna(desc) else ""
            status_str = str(status).strip() if pd.notna(status) else "Pending"
            
            # Handle float value
            try:
                val_num = float(value) if pd.notna(value) and str(value).strip() != '' else 0.0
            except (ValueError, TypeError):
                # Fallback to float if it contains commas or other numeric strings
                val_str = str(value).replace(',', '').strip()
                try:
                    val_num = float(val_str)
                except:
                    val_num = 0.0
            
            if isinstance(date_val, datetime.datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            elif pd.notna(date_val):
                date_str = str(date_val).split(' ')[0]
            else:
                date_str = "N/A"
                
            record = {
                "eq_no": eq_no,
                "description": desc_str,
                "date": date_str,
                "from_field": from_str,
                "status": status_str,
                "value": val_num
            }
            
            # If there are duplicates, keep the one with a more complete description or later date
            if digits in enquiries:
                prev = enquiries[digits]
                # Compare value or date
                if record["date"] != "N/A" and (prev["date"] == "N/A" or record["date"] > prev["date"]):
                    enquiries[digits] = record
            else:
                enquiries[digits] = record
                
        return enquiries
    except Exception as e:
        print(f"Error loading enquiries: {e}")
        return {}

def get_directory_cache():
    """Load folder cache from file or build it if missing."""
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return build_directory_cache()

def load_outlook_db():
    """Load communication records from JSON database."""
    if os.path.exists(OUTLOOK_DB_FILE):
        try:
            with open(OUTLOOK_DB_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading outlook db: {e}")
    return []

def get_verified_details(customer_name, project_id=None):
    """Return the correct verified customer name and location based on corporate directory."""
    cust_upper = customer_name.upper() if customer_name else ""
    
    if "SOMAFUSION" in cust_upper:
        return "SOMAFUSION WELLNESS LLP", "Mumbai"
    elif "ENERTICS" in cust_upper:
        return "ENERTICS", "Panchkula"
    elif "AEBOCODE" in cust_upper:
        return "Aebocode Technologies", "Panchkula"
    elif "ORIENT" in cust_upper:
        return "ORIENT", "Noida"
    elif "GOEL" in cust_upper:
        return "GOEL ENTERPRISES SCANNING", "Panchkula"
    elif "HELLA" in cust_upper:
        return "HELLA", "Dera Bassi"
    elif "HENKEL" in cust_upper:
        if "4 SETS" in cust_upper:
            return "Henkel 4 sets", "Baddi"
        elif "2 SETS" in cust_upper:
            return "HENKEL 2 SETS REPEAT", "Baddi"
        return "HENKEL", "Baddi"
    elif "AV MACHINING" in cust_upper:
        return "AV MACHINING SOLUTIONS", "Mohali"
    elif "BAJAJ" in cust_upper:
        return "BAJAJ", "Pune"
    elif "SONALIKA" in cust_upper:
        if "REPEAT ORDER 12 NOS" in cust_upper:
            return "Sonalika RepeAT Order 12 Nos", "Hoshiarpur"
        return "SONALIKA", "Hoshiarpur"
    elif "ENDUREAIR" in cust_upper:
        return "ENDUREAIR", "Noida"
    elif "3D PARADISE" in cust_upper:
        return "3D PARADISE", "Ludhiana"
    elif "HC ROBOTICS" in cust_upper:
        return "HC ROBOTICS", "Hyderabad"
    elif "MENTHOSA" in cust_upper:
        return "MENTHOSA", "Ludhiana"
    elif "NIPA" in cust_upper:
        return "NIPA", "Mohali"
    elif "LABAT" in cust_upper:
        if "REPEAT" in cust_upper:
            return "LABAT REPEAT VC ABS", "Hoshiarpur"
        elif "ASIA" in cust_upper:
            return "LABAT ASIA PUR / 2025-26 / WOT-46", "Hoshiarpur"
        return "LABAT ASIA", "Hoshiarpur"
    elif "PARAM" in cust_upper:
        return "PARAM APPLIANCES", "Ludhiana"
    elif "KYOCERA" in cust_upper:
        return "KYOCERA", "Gurgaon"
    elif "LOTUS" in cust_upper:
        return "LOTUS DEFENCE", "Panchkula"
    elif "SG ENGINEERING" in cust_upper:
        return "SG ENGINEERING WORKS", "Hoshiarpur"
    elif "UNO MINDA" in cust_upper:
        return "UNO MINDA", "Gurgaon"
    elif "JATIN" in cust_upper or "SHINDE" in cust_upper:
        return "ORIENT", "Noida"
    elif "UNDER PROCESS" in cust_upper:
        return "UNO MINDA", "Gurgaon"
        
    return customer_name, None

def resolve_project_details(project_id, cache, enquiries, outlook_db=None):
    """Combine data from master list, folder cache, and outlook db to get full details."""
    if outlook_db is None:
        outlook_db = load_outlook_db()
        
    digits_match = re.search(r'\d+', str(project_id))
    if not digits_match:
        return None
    digits = digits_match.group(0)
    
    # Retrieve records
    enq_rec = enquiries.get(digits, {})
    cache_rec = cache.get(digits, {})
    
    # Find matching communication record
    comm_rec = {}
    for entry in outlook_db:
        ref_no = entry.get("referenceNo", "")
        proj_code = entry.get("projectCode", "")
        if proj_code == project_id or (digits in str(ref_no)) or (digits in str(proj_code)):
            comm_rec = entry
            break
            
    # Merge rules
    # Customer name: Prefer comm_rec client, fallback to enquiries FROM, fallback to client folder name, fallback to description if it is short, fallback to "Unknown Customer"
    from_f = enq_rec.get("from_field", "")
    folder_client = cache_rec.get("client_folder_name", "")
    desc = enq_rec.get("description", "")
    comm_client = comm_rec.get("client", "")
    
    customer_name = "N/A"
    if comm_client:
        customer_name = comm_client
    elif from_f:
        customer_name = from_f
    elif folder_client:
        customer_name = folder_client
    elif desc and len(desc) < 40:
        customer_name = desc
        
    company_name = customer_name
    
    # Site details & address
    address = comm_rec.get("address", cache_rec.get("address", "N/A"))
    location = comm_rec.get("city", cache_rec.get("city", "N/A"))
    
    # Apply verified details overrides
    lookup_name = customer_name if customer_name != "N/A" else desc
    verified_cust, verified_loc = get_verified_details(lookup_name, project_id)
    if verified_cust and verified_cust != "N/A":
        customer_name = verified_cust
        company_name = verified_cust
    if verified_loc and verified_loc != "N/A":
        location = verified_loc
        
    # Project Status
    status = enq_rec.get("status", "Pending")
    if pd.isna(status) or status == "nan" or not status:
        status = "In-Orbit"
        
    # Project Value
    value = enq_rec.get("value", 0.0)
    
    merged = {
        "project_id": f"C{digits}",
        "eq_no": enq_rec.get("eq_no", f"EQ{digits}"),
        "customer_name": customer_name,
        "company_name": company_name,
        "description": desc or "N/A",
        "date": enq_rec.get("date", "N/A"),
        "site_name": "Englabs Project Site",
        "site_address": address,
        "location": location,
        "city": location if location != "N/A" else comm_rec.get("city", cache_rec.get("city", "N/A")),
        "state": comm_rec.get("state", cache_rec.get("state", "N/A")),
        "contact_person": comm_rec.get("contactPerson", cache_rec.get("contact_person", "N/A")),
        "mobile_number": comm_rec.get("mobileNumber", cache_rec.get("mobile_number", "N/A")),
        "email_id": comm_rec.get("email", "N/A"),
        "latest_communication_date": comm_rec.get("date", "N/A"),
        "project_status": status,
        "project_value": value
    }
    
    return merged

def backup_site_cash():
    """Create a backup of Englabs Site Cash.xlsx."""
    if not os.path.exists(SITE_CASH_FILE):
        return None
    os.makedirs(BACKUP_DIR, exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = os.path.join(BACKUP_DIR, f"Englabs Site Cash_Backup_{timestamp}.xlsx")
    shutil.copy2(SITE_CASH_FILE, backup_file)
    print(f"Backup created at: {backup_file}")
    return backup_file

def log_update_history(backup_path, updated_records, action="reconcile"):
    """Append update summary to the JSON history log file."""
    history = []
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                history = json.load(f)
        except:
            pass
            
    log_entry = {
        "timestamp": datetime.datetime.now().isoformat(),
        "action": action,
        "backup_path": backup_path,
        "updated_count": len(updated_records),
        "records": updated_records
    }
    history.insert(0, log_entry) # Put newest first
    
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(history, f, indent=4, ensure_ascii=False)

def reconcile_all_excel(specific_ids=None):
    """Safely reconcile details in the Englabs Site Cash.xlsx ledger."""
    if not os.path.exists(SITE_CASH_FILE):
        return {"success": False, "error": f"Site cash file not found: {SITE_CASH_FILE}"}
        
    try:
        # Load sheets
        cache = get_directory_cache()
        enquiries = load_enquiries()
        outlook_db = load_outlook_db()
        
        # Open with openpyxl to preserve formulas/styles
        wb = openpyxl.load_workbook(SITE_CASH_FILE, data_only=False)
        sheet = wb['Projects wise Details_2026']
        
        # Style definition for professional look
        from openpyxl.styles import Font, Alignment, Border, Side
        cambria_font = Font(name="Cambria", size=10, bold=False)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        
        # Create backup
        backup_path = backup_site_cash()
        
        updated_records = []
        unmatched_ids = []
        missing_details_ids = []
        duplicate_ids_count = 0
        total_rows_processed = 0
        
        # In openpyxl, rows are 1-indexed.
        # Row 1 is title, Row 2 is headers, data starts at Row 3.
        max_row = sheet.max_row
        
        # Track IDs to find duplicates in Site Cash
        site_cash_ids = []
        
        for r in range(3, max_row + 1):
            project_id_cell = sheet.cell(row=r, column=2) # Column B
            original_val = project_id_cell.value
            project_id = str(original_val).strip() if original_val is not None else ""
            if not project_id or project_id == 'nan' or project_id == 'Current Projects Detasils (Project ID)':
                continue
            if project_id == 'TOTAL':
                continue
                
            total_rows_processed += 1
            
            # If specific_ids filter is provided, check it
            if specific_ids and project_id not in specific_ids:
                continue
                
            # Track duplicates
            if project_id in site_cash_ids:
                duplicate_ids_count += 1
            else:
                site_cash_ids.append(project_id)
                
            digits_match = re.search(r'\d+', project_id)
            if not digits_match:
                unmatched_ids.append(project_id)
                continue
            digits = digits_match.group(0)
            
            # Lookup
            details = resolve_project_details(project_id, cache, enquiries, outlook_db)
            if not details:
                unmatched_ids.append(project_id)
                continue
                
            # Determine Customer Name and Location
            cust_name = details["customer_name"]
            loc = details["location"]
            
            # Override with verified mappings if possible
            lookup_name = cust_name if cust_name != "N/A" else details.get("description", "")
            verified_cust, verified_loc = get_verified_details(lookup_name, project_id)
            if verified_cust and verified_cust != "N/A":
                cust_name = verified_cust
            if verified_loc and verified_loc != "N/A":
                loc = verified_loc
                
            # If details are missing
            if cust_name == "N/A" or loc == "N/A":
                missing_details_ids.append(project_id)
                
            # Safely write to Columns C (3) and D (4)
            cust_cell = sheet.cell(row=r, column=3)
            loc_cell = sheet.cell(row=r, column=4)
            
            curr_cust = str(cust_cell.value).strip() if cust_cell.value is not None else ""
            curr_loc = str(loc_cell.value).strip() if loc_cell.value is not None else ""
            
            # Clean up the Project ID cell styling and value (trim whitespace)
            cell_changed = False
            if original_val != project_id:
                project_id_cell.value = project_id
                project_id_cell.font = cambria_font
                project_id_cell.alignment = center_align
                project_id_cell.border = thin_border
                cell_changed = True
                
            updated = False
            
            # Always overwrite if different and not N/A
            if cust_name != "N/A" and curr_cust != cust_name:
                cust_cell.value = cust_name
                cust_cell.font = cambria_font
                cust_cell.alignment = center_align
                cust_cell.border = thin_border
                updated = True
                
            if loc != "N/A" and curr_loc != loc:
                loc_cell.value = loc
                loc_cell.font = cambria_font
                loc_cell.alignment = center_align
                loc_cell.border = thin_border
                updated = True
                
            if updated or cell_changed:
                updated_records.append({
                    "row": r,
                    "project_id": project_id,
                    "customer_name": cust_name,
                    "location": loc
                })
                
        # Autofit column dimensions professionally for columns A to N
        for col_idx in range(1, 15):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_len = 0
            for cell in sheet[col_letter]:
                # Calculate len, ignore formulas (starting with '=') and None
                if cell.value and not str(cell.value).startswith('='):
                    max_len = max(max_len, len(str(cell.value)))
            # Standard padding
            sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)
            
        # Save workbook if updates were made
        if updated_records:
            wb.save(SITE_CASH_FILE)
            log_update_history(backup_path, updated_records, action="reconcile")
            print(f"Excel updated successfully. {len(updated_records)} rows filled.")
        else:
            print("No new updates to write to Excel (cells already populated or no matches).")
            
        return {
            "success": True,
            "updated_count": len(updated_records),
            "total_processed": total_rows_processed,
            "duplicate_count": duplicate_ids_count,
            "unmatched_ids": unmatched_ids,
            "missing_details_ids": missing_details_ids,
            "updated_records": updated_records,
            "backup_path": backup_path
        }
    except Exception as e:
        print(f"Error during reconciliation: {e}")
        return {"success": False, "error": str(e)}

def detect_ids_from_image(image_path):
    """Use OCR to extract project IDs from a screenshot image."""
    if not os.path.exists(image_path):
        return []
    try:
        img = Image.open(image_path)
        text = pytesseract.image_to_string(img)
        # Find CXXXX or EQXXXX
        matches = re.findall(r'(?:C|EQ)\s*(\d{4,5})', text, re.IGNORECASE)
        # Clean spacing
        cleaned = [f"C{m}" for m in matches]
        # Also look for raw 4-to-5 digit numbers as fallbacks
        raw_numbers = re.findall(r'\b\d{4,5}\b', text)
        for num in raw_numbers:
            # check if it looks like a valid enquiry digit
            cleaned.append(f"C{num}")
            
        return list(set(cleaned))
    except Exception as e:
        print(f"OCR Error: {e}")
        return []

def detect_ids_from_pdf(pdf_path):
    """Parse text from PDF file to extract project IDs."""
    if not os.path.exists(pdf_path):
        return []
    try:
        doc = fitz.open(pdf_path)
        text = ""
        for page in doc:
            text += page.get_text()
            
        matches = re.findall(r'(?:C|EQ)\s*(\d{4,5})', text, re.IGNORECASE)
        cleaned = [f"C{m}" for m in matches]
        raw_numbers = re.findall(r'\b\d{4,5}\b', text)
        for num in raw_numbers:
            cleaned.append(f"C{num}")
            
        return list(set(cleaned))
    except Exception as e:
        print(f"PDF Parse Error: {e}")
        return []

def detect_ids_from_excel(excel_path):
    """Read all sheets and cell values in Excel to extract project IDs."""
    if not os.path.exists(excel_path):
        return []
    try:
        xl = pd.ExcelFile(excel_path)
        found_ids = []
        for s in xl.sheet_names:
            df = xl.parse(s, header=None)
            for col in df.columns:
                # String conversion and regex match
                col_strs = df[col].astype(str)
                for val in col_strs:
                    matches = re.findall(r'(?:C|EQ)\s*(\d{4,5})', val, re.IGNORECASE)
                    for m in matches:
                        found_ids.append(f"C{m}")
        return list(set(found_ids))
    except Exception as e:
        print(f"Excel Read Error: {e}")
        return []

def process_file_detection(file_path):
    """Detect format and extract details."""
    ext = os.path.splitext(file_path)[1].lower()
    detected_ids = []
    
    if ext in ['.jpg', '.jpeg', '.png', '.tiff', '.bmp']:
        detected_ids = detect_ids_from_image(file_path)
    elif ext == '.pdf':
        detected_ids = detect_ids_from_pdf(file_path)
    elif ext in ['.xlsx', '.xls']:
        detected_ids = detect_ids_from_excel(file_path)
        
    print(f"Detected project IDs from {os.path.basename(file_path)}: {detected_ids}")
    
    # Look up details for all detected IDs
    cache = get_directory_cache()
    enquiries = load_enquiries()
    
    results = []
    for pid in detected_ids:
        details = resolve_project_details(pid, cache, enquiries)
        if details:
            results.append(details)
            
    return {
        "detected_ids": detected_ids,
        "results": results
    }

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        cmd = sys.argv[1]
        if cmd == "--reconcile-all":
            res = reconcile_all_excel()
            print(json.dumps(res, indent=4))
        elif cmd == "--reconcile-ids" and len(sys.argv) > 2:
            ids = sys.argv[2:]
            res = reconcile_all_excel(specific_ids=ids)
            print(json.dumps(res, indent=4))
        elif cmd == "--detect-file" and len(sys.argv) > 2:
            path = sys.argv[2]
            res = process_file_detection(path)
            print(json.dumps(res, indent=4))
        elif cmd == "--build-cache":
            build_directory_cache()
        else:
            print("Unknown command arguments")
    else:
        # Default: print status of a single project lookup for testing
        cache = get_directory_cache()
        enquiries = load_enquiries()
        test_details = resolve_project_details("C4465", cache, enquiries)
        print("Test Lookup C4465:", json.dumps(test_details, indent=4))
