import openpyxl
from datetime import datetime

file_path = r"G:\HR Team Managements\Englabs Projects APK\Projects\Current Projects  DETAISL.xlsx"
projects_to_add = [
    {'Project ID': 'EQ5487', 'Client Name': 'EndureAir', 'Status': 'Pending', 'Date Added': datetime.now().strftime('%Y-%m-%d')},
    {'Project ID': 'EQ5488', 'Client Name': 'RND Mechanical', 'Status': 'Pending', 'Date Added': datetime.now().strftime('%Y-%m-%d')},
    {'Project ID': 'EQ5489', 'Client Name': 'Aebocode', 'Status': 'Pending', 'Date Added': datetime.now().strftime('%Y-%m-%d')},
    {'Project ID': 'EQ5490', 'Client Name': 'Orient Electric 1', 'Status': 'Pending', 'Date Added': datetime.now().strftime('%Y-%m-%d')},
    {'Project ID': 'EQ5491', 'Client Name': 'Orient Electric 2', 'Status': 'Pending', 'Date Added': datetime.now().strftime('%Y-%m-%d')}
]

try:
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # Find true last row by looking for last Customer Name (Column C)
    last_row = 4
    last_sr_no = 0
    for row_idx in range(5, sheet.max_row + 10):
        cust_val = sheet.cell(row=row_idx, column=3).value
        sr_val = sheet.cell(row=row_idx, column=2).value
        
        if cust_val:
            last_row = row_idx
            if isinstance(sr_val, (int, float)):
                last_sr_no = int(sr_val)
        
    start_row = last_row + 1
    
    print(f"Detected last row at {last_row}. Last Sr No: {last_sr_no}. Appending at row {start_row}...")
    
    for i, proj in enumerate(projects_to_add):
        current_row = start_row + i
        
        # B: sr. no.
        sheet.cell(row=current_row, column=2).value = last_sr_no + 1 + i
        
        # C: Customer Name
        sheet.cell(row=current_row, column=3).value = proj['Client Name']
        
        # J: Project id
        sheet.cell(row=current_row, column=10).value = proj['Project ID']
        
        # R: Status
        sheet.cell(row=current_row, column=18).value = proj['Status']
        
        # E: Dated
        sheet.cell(row=current_row, column=5).value = proj['Date Added']

    wb.save(file_path)
    print(f"Successfully appended {len(projects_to_add)} projects.")

except Exception as e:
    print(f"Error updating Excel: {e}")
